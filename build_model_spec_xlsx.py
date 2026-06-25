"""Generate K_2_Cents_Model_Spec.xlsx — 8-sheet model specification.

Pulls live data from the recommender output so the workbook reflects the
current model state, not a stale snapshot.
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
PROC = ROOT / "data" / "processed"
OUT = ROOT / "K_2_Cents_Model_Spec.xlsx"

FONT = "Arial"
H1 = Font(name=FONT, size=14, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=11, bold=True)
H_FILL = PatternFill("solid", start_color="1F4E78")
SUB_FILL = PatternFill("solid", start_color="DDEBF7")
NOTE_FILL = PatternFill("solid", start_color="FFF2CC")
BODY = Font(name=FONT, size=10)
BODY_BOLD = Font(name=FONT, size=10, bold=True)
THIN = Side(border_style="thin", color="B7B7B7")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(sheet, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = sheet.cell(row=row, column=c)
        cell.font = H1
        cell.fill = H_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(sheet, start_row: int, end_row: int, cols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = sheet.cell(row=r, column=c)
            if cell.font.name != FONT:
                cell.font = BODY
            cell.alignment = WRAP
            cell.border = BORDER


def set_widths(sheet, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = w


def add_title(sheet, title: str, subtitle: str = "") -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(name=FONT, size=16, bold=True, color="1F4E78")
    if subtitle:
        sheet["A2"] = subtitle
        sheet["A2"].font = Font(name=FONT, size=10, italic=True, color="595959")


# ─── Sheet 1: Overview ───────────────────────────────────────────────────────


def sheet_overview(wb: Workbook) -> None:
    s = wb.active
    s.title = "Overview"
    add_title(s, "K's 2 Cents — Fantasy Recommender Model Specification",
              f"Built: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Anchored to FIFA Fantasy WC2026 scoring schema")
    s["A4"] = "MODEL FLOW"
    s["A4"].font = H2
    flow = [
        ["Stage", "Input", "Process", "Output"],
        ["1. Fixture profile (§A)",
         "Polymarket markets + 365 trends + weather + nation strength composite",
         "Per-fixture: goals_index, cs_index, fixture_shape ∈ {consensus_lopsided, consensus_tight, market_overconfident, composite_overconfident}",
         "24 rows per round"],
        ["2. Per-player scoring (§B, §C, §D)",
         "stg_players_view (165 cols), fantasy_player_totals, recent windows, %selected",
         "Floor / Ceiling / Differential in 3 normalization modes (p90 / per_app / totals)",
         "1248 player-fixture rows"],
        ["3. Archetype enrichment v2 (§C.5)",
         "Scoring-channel composition + FIFA p90 stats + static profile",
         "K-means clustering by position class; retrospective (MD1+MD2 top-20%) + prospective (full pool)",
         "peer_archetype + similarity + top-3 exemplars per row"],
        ["4. Filters (§E)",
         "is_active, injury, eliminated flags",
         "Drop rows; tag anti_picks on consensus_tight fixtures",
         "Filtered recommendation set"],
        ["5. Position Suggestor",
         "Filtered recs + reason_chips",
         "Top 15 by ev_ensemble (ranked) + Look-out-for: 5 DEF + 5 MID + 5 FWD + 2 GK by value_score",
         "wc26_fantasy_position_suggestor.json"],
        ["6. Strategy squads (§D)",
         "Filtered recs + strategy weights (α, β, γ) + SB-quota",
         "Greedy 15-man assembly with hard constraints: 2/5/5/3 quota, ≤3 per nation, ≤$100m, SB-band cap",
         "wc26_fantasy_strategy_squads.json (3 squads + 12th-man each)"],
    ]
    for i, row in enumerate(flow, start=5):
        for j, val in enumerate(row, start=1):
            s.cell(row=i, column=j, value=val)
    style_header_row(s, 5, 4)
    style_body(s, 6, 5 + len(flow) - 1, 4)
    set_widths(s, {1: 30, 2: 40, 3: 60, 4: 28})

    # Architecture note
    note_r = 5 + len(flow) + 1
    s.cell(row=note_r, column=1, value="ARCHITECTURE").font = H2
    s.cell(row=note_r + 1, column=1,
           value="All compute lives in the warehouse (E:/fc2026_notebooks/lib/recommender.py + 17_fantasy_recommender.py). "
                 "The PWA only renders. Hourly tick produces wc26_fantasy_recommendations.json, "
                 "wc26_fantasy_position_suggestor.json, wc26_fantasy_strategy_squads.json. "
                 "Snapshot ts on every row — slider re-filters from snapshot, never re-fetches mid-round.")
    s.cell(row=note_r + 1, column=1).alignment = WRAP
    s.cell(row=note_r + 1, column=1).fill = NOTE_FILL
    s.merge_cells(start_row=note_r + 1, end_row=note_r + 2, start_column=1, end_column=4)


# ─── Sheet 2: Factor catalog ─────────────────────────────────────────────────


def sheet_factor_catalog(wb: Workbook) -> None:
    s = wb.create_sheet("Factor catalog")
    add_title(s, "Factor catalog — every §A through §I factor",
              "EDA |corr| col = MAX |corr| between the factor (computed over MD1+MD2) "
              "and actual MD1+MD2 fantasy points. "
              "CAVEAT: this is CONTEMPORANEOUS correlation — not predictive validation. "
              "Factors that ARE the events points are computed from (goals, assists, "
              "clean sheets, form, avg_points) will show inflated |corr| by construction. "
              "See AUTOCORRELATION AUDIT below — autocorrelated factors are flagged "
              "with handling notes in the Definition column. "
              "'Status' = ACTIVE / DROPPED / DEFERRED.")

    # Autocorrelation audit — three-row block above the headers so readers
    # see the methodology caveat BEFORE scanning the |corr| column.
    s["A3"] = "AUTOCORRELATION AUDIT"
    s["A3"].font = H2
    audit = [
        ["Question", "Answer"],
        ["Why does |corr| look so high for some factors?",
         "Fantasy points are LITERALLY a function of in-match events: goal=4-6pts, assist=3pts, clean_sheet=4/5pts, save_chunks=+1, etc. Any factor that IS one of those event counts (B7 shots_on_target=0.65, B11 CS_prior=0.44, C1 XG=0.63, C2 assists=0.42) trivially correlates with points because it's almost computed from the same data. Same trap for LAGGED-TARGET factors (form, avg_pts, total_pts) — they ARE past fantasy points being used to predict future fantasy points."],
        ["How is this handled in the model?",
         "Three mechanisms: (1) ROUND-LOCK isolates 'lagged' from 'concurrent' — R3 emits only see ≤R2 stats, so B4 Fantasy Meta becomes a genuinely lagged signal, not a sneak preview of R3 itself. (2) DAMPENERS — C-NEW avg_points carries a hard ×0.3 coefficient; form prefers independent fotmob_rating (0.46) over fantasy.form (0.80). (3) ENSEMBLE DIVERSIFICATION — M3 Stat Maximizer holds w_B4 at 0.10 (vs 0.35-0.40 in M1/M4) so the consensus signal isn't dominated by lagged points."],
        ["Which factors are autocorrelated with the target?",
         "HIGH risk (lagged target itself): C5 form 0.80 / C-NEW avg_points 0.83 / Bracket B4 (uses form, avg_pts, total_pts, last_round_points). MEDIUM (event-coupled — same window): B7 SoT 0.65 / B11 CS_prior 0.44 / C1 XG 0.63 / B14 powerrank 0.63. LOW (independent / pre-tournament priors): A14 nation_strength 0.71 / A2b moneyline_lopsidedness / A4 goals_index / B1 start_prob 0.17 / A12 365_trends. Independent signals are the model's predictive backbone; event-coupled features are gated by the round-lock so they're lagged not concurrent."],
        ["Is M3 Stat Maximizer the 'autocorrelation-free' counter-weight?",
         "Yes — by design. M3 weights B2 WC Perf at 0.50 (per-match FIFA stats — event-coupled but lagged via lock) and B4 Fantasy Meta at just 0.10. Its disagreement with M1 Banker (w_B4=0.35) is what makes consensus picks meaningful — when M3 AND M1 agree, the player is good both by independent stats and by lagged fantasy meta. Surprises from M3 alone tend to be undervalued by the fantasy market because they're not yet showing up in pts/form."],
    ]
    for i, row in enumerate(audit, start=4):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 4, 2)
    style_body(s, 5, 4 + len(audit) - 1, 2)
    # Spacer row before main table
    main_row_start = 4 + len(audit) + 2
    s.cell(row=main_row_start - 1, column=1, value="FACTOR INVENTORY").font = H2

    headers = ["ID", "Name", "Bucket", "Position scope", "Source table.column",
               "Transform", "EDA |corr|", "Weight (default)", "Status", "Definition / notes"]
    rows = [
        # §A Fixture
        ["A1", "Stage flip multiplier", "Fixture", "All", "wc26_stg_matches.stage",
         "group=1.00, R32=1.15, R16=1.25, QF=1.40, SF+=1.55", "—", "medium", "ACTIVE",
         "Higher stakes → weaker teams overperform; multiplies fixture indices."],
        ["A2a", "Moneyline volume (consensus strength)", "Fixture", "All",
         "wc26_polymarket_match_volume.volume_moneyline",
         "Percentile rank → 0..1 confidence weight on A2b", "—", "light", "ACTIVE",
         "Gates A2b's confidence — low volume = soft signal."],
        ["A2b", "Moneyline lopsidedness", "Fixture", "All",
         "wc26_match_polymarket_markets (moneyline)", "p_home_win − p_away_win", "—", "HEAVY", "ACTIVE",
         "Signed [-1, +1]; positive = home favored. Drives consensus_lopsided shape."],
        ["A2c", "Other-market depth", "Fixture", "All", "—", "—", "0.09", "drop", "DROPPED",
         "Corr 0.09 with goal totals on MD1+MD2 — no signal."],
        ["A3", "Per-side clean-sheet probability", "Fixture", "DEF, GK, MID",
         "wc26_match_polymarket_markets (team-total-{home|away}-0pt5)",
         "p_side_cs = 1 - p(opp scores 0.5+)", "—", "HEAVY", "ACTIVE",
         "Per-side derivation from O/U 0.5 Yes-price markets."],
        ["A4", "Over/under 2.5 goals (goals_index)", "Fixture", "All",
         "wc26_match_polymarket_markets (total-2pt5)", "p(over 2.5) × weather mod × roof mod",
         "—", "HEAVY", "ACTIVE", "Drives goal_prob ceiling component."],
        ["A5", "Per-player goalscorer odds", "Fixture/Player", "FWD, MID",
         "wc26_match_polymarket_markets (anytime_goalscorer)",
         "Direct implied prob; null fallback to fifa_wc_XG", "—", "medium", "ACTIVE (sparse)",
         "Polymarket coverage thin — only 16/104 matches; fallback dominates."],
        ["A6", "Referee discipline", "Fixture", "All", "wc26_stg_referee_profile.yellow_pg",
         "Normalize vs panel median → card_index 0..1", "—", "light", "ACTIVE",
         "Multiplied with C-family attackers in high-card fixtures."],
        ["A7", "Weather cluster", "Fixture", "All",
         "wc26_match_weather + wc26_stg_stadiums.surface/roof_type",
         "4 EDA clusters: 0 mild, 1 hot/dry, 2 altitude, 3 wet. Modifiers in D-6.",
         "0.27 (cluster 1 draw rate)", "medium", "ACTIVE",
         "Cluster 1: goals×0.92, cs×1.10. Cluster 2: goals×0.88. Retractable roof: goals×1.15."],
        ["A8", "Stadium altitude", "Fixture", "All", "—", "—", "—", "drop", "DROPPED",
         "Subsumed into weather cluster 2 (altitude venues)."],
        ["A9", "FIFA-rank gap", "Fixture", "All", "wc26_stg_nations.fifa_rank",
         "home_rank − away_rank, bucketed", "0.71 (nation_strength as proxy)", "medium", "ACTIVE",
         "Underlies §I.1 static profile + A14."],
        ["A10", "Nation tournament form", "Fixture", "All",
         "Derived from fantasy_round_matches prior rounds",
         "W=3 D=1 L=0 over up-to-2 matches", "—", "medium", "ACTIVE",
         "Folded into §I.2 nation_total_strength."],
        ["A11", "WC champion implied prob", "Fixture", "All",
         "wc26_polymarket_winner_history (latest snapshot)",
         "Direct 0..1", "—", "light", "ACTIVE",
         "Long-term sentiment signal beyond this fixture."],
        ["A12", "365scores top trend", "Fixture", "All",
         "wc26_stg_match_trends_365",
         "Category × pct-bucket calibration multiplier (EDA §4)", "0.62 baseline",
         "HEAVY (when category matches)", "ACTIVE",
         "doubleChance≥0.9: 95% hit (n=21). BTTS≥0.9: 100% (n=4). Result-type: 44-75% discount."],
        ["A13", "365scores decider trend", "Fixture", "All",
         "wc26_stg_match_trends_365 (confidenceTrendIds)",
         "Same calibration as A12", "—", "medium", "ACTIVE",
         "Confirmation strength back-up signal."],
        ["A14", "Nation strength delta", "Fixture", "All",
         "§I composite (I.1 + I.2 + I.4)", "Stage-conditional weights; D-2",
         "0.71", "HEAVY", "ACTIVE",
         "Independent of Polymarket; signed [-1, +1] home perspective."],
        ["A15", "Market-vs-composite divergence", "Fixture", "All",
         "|A2b - A14|", "Abs value → upset opportunity flag", "—", "medium", "ACTIVE",
         "High divergence → market_overconfident or composite_overconfident shape."],
        # §B Floor
        ["B1", "Start probability", "Floor", "All",
         "recent5_started_pct → recent10_started_pct → 0.5",
         "Direct 0..1, fallback cascade", "0.17", "HEAVY", "ACTIVE",
         "Gates appearance points + every other position contribution."],
        ["B3", "GK saves rate × opp goals_index", "Floor", "GK",
         "fifa_wc_GoalkeeperSaves / appearances",
         "Per-mode denom × 1/3 (every 3 saves = +1)", "0.21", "medium", "ACTIVE",
         "Maps to FIFA's every-3-saves +1 schema."],
        ["B4", "DEF tackles per 90", "Floor", "DEF", "fantasy_totals.tackles",
         "Per-mode denom (no points-bonus for DEF tackles)", "0.05", "light", "ACTIVE",
         "Activity proxy; helps CS prediction indirectly."],
        ["B5", "MID tackle volume", "Floor", "MID", "fantasy_totals.tackles",
         "Per-mode denom × 1/3 (FIFA: every 3 tackles → +1)", "0.17", "medium", "ACTIVE",
         "Reframed from binary floor (was 4.5/app) to ceiling-contributor."],
        ["B6", "MID chances-created volume", "Floor", "MID", "fantasy_totals.chances_created",
         "Per-mode denom × 1/2 (every 2 CC → +1)", "0.39", "medium", "ACTIVE",
         "Same reframe — volume contributor, not floor gate."],
        ["B7", "FWD shots-on-target volume", "Floor", "FWD",
         "fifa_wc_AttemptAtGoalOnTarget OR fantasy_totals.shots_on_target",
         "Per-mode denom × 1/2 (every 2 SoT → +1)", "0.65", "HEAVY", "ACTIVE",
         "Strong predictor of FWD output."],
        ["B11", "DEF clean-sheet prior", "Floor", "DEF",
         "fifa_wc_CleanSheets / appearances",
         "× A3 (fixture per-side CS prob) × 5 pts", "0.44", "HEAVY", "ACTIVE",
         "Maps to FIFA's DEF CS bonus (+5)."],
        ["B14", "Power-rank within team", "Floor", "All",
         "wc26_stg_player_powerrank.avg_*_score",
         "Position-conditional: atk→FWD, def→DEF, cre→MID, gk→GK", "0.63", "HEAVY", "ACTIVE",
         "Strong cross-position signal."],
        # §C Ceiling
        ["C1", "Goal probability", "Ceiling", "FWD, MID",
         "fifa_wc_XG / denom + recent5/10/15_goals decay",
         "Decay weights {0.5, 0.3, 0.2} × goals_index × 2",
         "0.63 (fifa_wc_XG)", "HEAVY", "ACTIVE",
         "Combined per-mode XG and recent goal rate."],
        ["C2", "Assist probability", "Ceiling", "MID, FWD",
         "fifa_wc_Assists / denom + fotmob_wc_big_chances_created",
         "Combined index × goals_index × 3 pts", "0.42", "HEAVY", "ACTIVE",
         "Maps to +3 assist points."],
        ["C5", "Form multiplier", "Ceiling", "All",
         "recent5_fotmob_rating (independent) OR fantasy.form",
         "((rating - 6.0) / 4.0).clip(-0.15, 0.20) + 1.0 → multiplier 0.85-1.20",
         "0.46 (fotmob, independent); 0.80 (form, autocorr)", "medium", "ACTIVE",
         "Prefer fotmob rating — form is partly autocorrelated with points."],
        ["C7", "FotMob WC creativity", "Ceiling", "MID, FWD",
         "fotmob_wc_big_chances_created + fotmob_wc_touches_opp_box",
         "Combined index 0..1", "0.39", "medium", "ACTIVE", "Sub-component of C1/C2."],
        ["C-NEW", "avg_points lag-1 prior", "Ceiling/Floor", "All",
         "fantasy_players.avg_points", "× 0.3 dampener (autocorrelation)", "0.83", "medium", "ACTIVE",
         "Added from EDA §10 broad sweep; weighted low to avoid lock-in."],
        ["C-NEW", "recent5_player_of_the_match", "Ceiling", "All",
         "wc26_stg_players_view.recent5_player_of_the_match", "× 0.5 small bonus", "0.45", "light", "ACTIVE",
         "Added from EDA §10."],
        # §D Differential
        ["D1", "Live %selected (SB sigmoid)", "Differential", "All",
         "fantasy_players.percent_selected",
         "1 / (1 + exp((pct − 5) / 1)) — sigmoid gate around 5%", "0.47 (POSITIVE)", "HEAVY", "ACTIVE (reframed)",
         "Raw ownership correlates POSITIVELY with points (popular = quality). "
         "Differential value lives ONLY on the +2 SB event."],
        ["D2", "SB track record", "Differential", "All",
         "wc26_stg_fantasy_player_totals.scouting_bonus",
         "× (1 + 0.2 × min(sb_total, 5)) multiplier on D1", "0.68", "HEAVY", "ACTIVE",
         "Already-earned SBs flag who clears the 4-pt threshold."],
        ["D3", "Ownership trend", "Differential", "All",
         "fantasy_players.rounds_selected_json (Δ last 2 rounds)",
         "Rising fast → cap D1; falling → boost", "—", "medium", "DEFERRED",
         "Implementation: pending — needs to parse rounds_selected_json."],
        ["D4", "Editorial one-to-watch", "Differential", "All",
         "fantasy_players.one_to_watch", "+0.3 small lift", "—", "light", "DEFERRED",
         "Column not exposed in current snapshot — defaulting to False."],
        # §I Nation strength
        ["I.1", "Static profile", "§I (nation)", "All (fixture-level)",
         "trophies_won + fifa_rank + squad_valuation + is_host + group + confederation",
         "norm01 components, weighted", "—", "0.33 (group stage)", "ACTIVE",
         "Tapers across stages per D-2: group 0.33 → KO 0.10."],
        ["I.2", "Tournament metrics (per-match)", "§I (nation)", "All",
         "wc26_stg_team_match_metrics (24+ metrics aggregated)",
         "All divided by matches_played (no raw totals)", "—", "0.33 (group)", "ACTIVE",
         "Per-match normalization avoids fixture-count skew."],
        ["I.3", "Heavy hitter / clutch", "§I (nation)", "All",
         "upset_win_rate, knockout_overperformance, clutch_close_score_rate",
         "Multiplier centered at 1.0 (D-4)", "—", "1.0 (no-op)", "DEFERRED",
         "Backfill from hand-curated CSV in later phase."],
        ["I.4", "Player cumulative strength", "§I (nation)", "All",
         "wc26_stg_player_powerrank summed per nation",
         "Sanity check on I.2; weighted sum 0.3 atk + 0.3 def + 0.2 cre + 0.2 gk", "—",
         "0.34 (group) → 0.50 (KO)", "ACTIVE",
         "Cross-check: if I.2/I.4 disagree → 'individual talent but poor cohesion'."],
        ["I.5", "Nation-vs-Nation delta", "§I (fixture)", "All",
         "home_total_strength − away_total_strength",
         "Normalized [-1, +1]; produces shape enum", "0.71", "HEAVY", "ACTIVE",
         "The §I output that feeds A14."],
    ]
    header_row = main_row_start
    body_start = header_row + 1
    for j, h in enumerate(headers, start=1):
        s.cell(row=header_row, column=j, value=h)
    style_header_row(s, header_row, len(headers))
    for i, row in enumerate(rows, start=body_start):
        for j, val in enumerate(row, start=1):
            s.cell(row=i, column=j, value=val)
    style_body(s, body_start, body_start + len(rows) - 1, len(headers))
    set_widths(s, {1: 6, 2: 30, 3: 12, 4: 14, 5: 35, 6: 35, 7: 16, 8: 14, 9: 14, 10: 60})
    s.freeze_panes = s.cell(row=body_start, column=1).coordinate


# ─── Sheet 3: Data lineage ───────────────────────────────────────────────────


def sheet_data_lineage(wb: Workbook) -> None:
    s = wb.create_sheet("Data lineage")
    add_title(s, "Data lineage — warehouse parquet → factor mapping",
              "Every active factor traced back to the source table + columns it consumes.")
    headers = ["Factor ID", "Source parquet", "Columns used", "Refresh cadence", "Notebook owner"]
    rows = [
        ["A1", "wc26_stg_matches", "stage", "Hourly", "15_staging_matches"],
        ["A2a/b", "wc26_match_polymarket_markets, wc26_polymarket_match_volume",
         "last_trade_price, outcome_prices, volume_moneyline, volume_other", "Hourly", "13_polymarket"],
        ["A3, A4, A5", "wc26_match_polymarket_markets (wide pivot)",
         "p_home_win, p_draw, p_over_*, p_btts*, p_*_cs (derived)", "Hourly", "13 + lib.recommender.build_match_markets_wide"],
        ["A6", "wc26_stg_referee_profile", "yellow_pg, red_pg (last_15)", "Daily", "04_referees → 14_staging_core"],
        ["A7", "wc26_match_weather + wc26_stg_stadiums",
         "temperature_c, humidity_pct, surface, roof_type", "Hourly + frozen",
         "12_match_weather + 02_stadiums"],
        ["A9", "wc26_stg_nations", "fifa_rank, squad_valuation_m_eur", "Frozen", "01_nations"],
        ["A11", "wc26_polymarket_winner_history", "latest_implied_prob per nation", "Hourly", "13_polymarket"],
        ["A12/A13", "wc26_match_trends_365",
         "lineTypeId, percentage, confidenceTrendIds, outcome (post-match)",
         "Hourly", "18_scores365_trends"],
        ["A14, A15", "wc26_stg_team_match_metrics + wc26_stg_player_powerrank + data/overrides/wc_trophies.csv",
         "§I composite inputs (40+ stats per team aggregation)", "Hourly + manual",
         "_build_stg_team_match_metrics.py + lib.recommender.nation_strength_composite"],
        ["B1, C5", "wc26_stg_players_view",
         "recent5_started_pct, recent10_started_pct, recent5_fotmob_rating, form",
         "Hourly", "16_staging_players"],
        ["B3-B7, C1-C2",
         "wc26_stg_players_view (FIFA WC totals) + wc26_stg_fantasy_player_totals",
         "fifa_wc_Goals/Assists/SoT/XG/CleanSheets/GoalkeeperSaves/TimePlayed, tackles, chances_created, shots_on_target",
         "Hourly", "16_staging_players + 10_fifa_fantasy"],
        ["B11, B14", "wc26_stg_player_powerrank",
         "avg_attacking_score, avg_defensive_score, avg_creativity_score, avg_defending_the_goal_score",
         "Hourly", "14_staging_core"],
        ["C-NEW (potm, recent goals)",
         "wc26_stg_players_view",
         "recent5_player_of_the_match, recent5_goals, recent10_goals, recent15_goals", "Hourly", "16_staging_players"],
        ["D1, D2", "fantasy_players + wc26_stg_fantasy_player_totals",
         "percent_selected, rounds_selected_json, scouting_bonus", "Hourly", "10_fifa_fantasy"],
        ["E1", "fantasy_players", "is_active", "Hourly", "10_fifa_fantasy"],
        ["§I.5 → A14", "lib/recommender.py::nation_strength_composite",
         "Builds per-nation index from I.1+I.2+I.4 weighted composite", "Hourly",
         "17_fantasy_recommender"],
        ["Top 15 + Look out for",
         "wc26_fantasy_recommendations.parquet",
         "ev_raw_{p90, per_app, totals}, percent_selected, reason_chips → build_position_suggestor",
         "Hourly", "17_fantasy_recommender"],
        ["3 strategy squads",
         "wc26_fantasy_recommendations.parquet + lib.recommender.STRATEGIES",
         "ev_strategy = α·floor + β·ceiling + γ·diff; greedy with constraints",
         "Hourly", "17_fantasy_recommender"],
        ["365 trends history", "wc26_match_polymarket_markets_history.parquet (D-5 append)",
         "Polymarket price snapshots with snapshot_ts for calibration retros", "Hourly", "13_polymarket"],
    ]
    for j, h in enumerate(headers, start=1):
        s.cell(row=4, column=j, value=h)
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            s.cell(row=i, column=j, value=val)
    style_header_row(s, 4, len(headers))
    style_body(s, 5, 5 + len(rows) - 1, len(headers))
    set_widths(s, {1: 20, 2: 45, 3: 60, 4: 16, 5: 30})
    s.freeze_panes = "A5"


# ─── Sheet 4: Scoring schema ─────────────────────────────────────────────────


def sheet_scoring(wb: Workbook) -> None:
    s = wb.create_sheet("Scoring schema")
    add_title(s, "FIFA Fantasy WC2026 — point values verbatim",
              "Source: in-app How to Score panel. Every model weight ties to these values.")

    s["A4"] = "All-players base"
    s["A4"].font = H2
    base = [
        ["Event", "Points"],
        ["Appearance (up to 60 min)", 1],
        ["Appearance (60+ min, cumulative)", 1],
        ["Assist", 3],
        ["Yellow card", -1],
        ["Red card", -2],
        ["Own goal", -2],
        ["Winning a penalty", 2],
        ["Conceding a penalty", -1],
    ]
    for i, row in enumerate(base, start=5):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 5, 2)
    style_body(s, 6, 5 + len(base) - 1, 2)

    s["A15"] = "Position-specific"
    s["A15"].font = H2
    pos_table = [
        ["Event", "GK", "DEF", "MID", "FWD"],
        ["Goal scored", 9, 7, 6, 5],
        ["Clean sheet (60+ min)", 5, 5, 1, "—"],
        ["First goal conceded", 0, 0, "—", "—"],
        ["Each additional GC", -1, -1, "—", "—"],
        ["Penalty save (excl. shootout)", 3, "—", "—", "—"],
        ["Every 3 saves", 1, "—", "—", "—"],
        ["Every 3 tackles", "—", "—", 1, "—"],
        ["Every 2 chances created", "—", "—", 1, "—"],
        ["Every 2 shots on target", "—", "—", "—", 1],
    ]
    for i, row in enumerate(pos_table, start=16):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 16, 5)
    style_body(s, 17, 16 + len(pos_table) - 1, 5)

    s["A27"] = "Bonuses"
    s["A27"].font = H2
    bonus = [
        ["Event", "Points"],
        ["Goal from direct free-kick", 1],
        ["Scouting Bonus: >4 pts in match AND in <5% of teams", 2],
    ]
    for i, row in enumerate(bonus, start=28):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 28, 2)
    style_body(s, 29, 28 + len(bonus) - 1, 2)

    s["A32"] = "Implications for the engine"
    s["A32"].font = H2
    impl = [
        "• DEF goals are extraordinarily valuable (7 pts vs FWD's 5).",
        "• GK appearance + CS + 1-save = +1+1+5+0 = 7 pts is the no-effort baseline for a starter on a favored side.",
        "• MID volume is granular but capped: 9 tackles = +3, 6 chances created = +3.",
        "• Scouting Bonus gate is <5%, not <8% (D-1-Doc had wrong threshold).",
        "• 60 min is NOT a hard threshold. A 45-min sub who assists scores +1+3 = 4 (Undav-type).",
        "• Captain bonus is ×2 actual points — captain choice IS part of EV, not post-hoc.",
        "• Penalty shootouts don't count (regulation + ET only).",
    ]
    for i, line in enumerate(impl, start=33):
        s.cell(row=i, column=1, value=line)
        s.cell(row=i, column=1).alignment = WRAP
        s.cell(row=i, column=1).fill = NOTE_FILL
        s.merge_cells(start_row=i, end_row=i, start_column=1, end_column=5)
    set_widths(s, {1: 40, 2: 12, 3: 10, 4: 10, 5: 10})


# ─── Sheet 5: Strategies ─────────────────────────────────────────────────────


def sheet_strategies(wb: Workbook) -> None:
    s = wb.create_sheet("Models")
    add_title(s, "Three independent models — one squad each (Challenges)",
              "Each model scores the same player pool with a different philosophy. The 3 squads track separately like real fantasy teams.")

    try:
        squads = json.loads((PROC / "wc26_fantasy_strategy_squads.json").read_text())
    except Exception:
        squads = []
    squad_by_id = {sq.get("model_id") or sq.get("strategy_id"): sq for sq in squads}

    MIDS = ["m1_banker", "m2_form_hunter", "m3_stat_max", "m4_sb_hunter"]

    headers = ["Attribute", "M1 — Banker", "M2 — Form Hunter", "M3 — Stat Maximizer", "M4 — SB Hunter"]
    rows = [
        ["Intent",
         "Floor-heavy. Consistency, average points, SB track record. Premium picks on favored fixtures. Cleanest top-decile path on average rounds.",
         "Recency-weighted. Recent goals, started-pct, fotmob rating, last-round explosions. Punishes cold streaks even on strong fixtures.",
         "Pure FIFA-stats. Powerrank + position-routed per-90 + creativity + duels. Ignores ownership entirely — picks best player regardless of crowd.",
         "Differential / SB-eligibility chaser. Custom assembler — anchors 3 fixed roles (2 'sure_shot_fwd' + 1 'influential_mid') then maximises sub-5% ownership across the remaining 12. High variance, high upside."],
        ["Bracket weight w_B1 (PlayerOverall)", 0.20, 0.10, 0.15, 0.15],
        ["Bracket weight w_B2 (WC Perf)",      0.25, 0.20, 0.50, 0.25],
        ["Bracket weight w_B3 (External)",     0.20, 0.40, 0.25, 0.20],
        ["Bracket weight w_B4 (Fantasy meta)", 0.35, 0.30, 0.10, 0.40],
        ["Post-boost sub-scores", "(none — bracket-only)",
         "recent_form_streak (recent5/10 goals, POM, started_pct, last_round_pts)",
         "creativity_engine (chances_created, big_chances, touches_opp_box, ball-progressions) + powerrank_pure (atk/def/cre/gk)",
         "sb_track_bonus (sb_total ≥1 multiplier, sigmoid gate at 5% ownership, differential_kicker for <2%)"],
        ["Fixture amplifier (B5 weight)", 1.00, 0.85, 1.00, 0.90],
        ["SB-quota (target, ±1)", "9 of 15", "6 of 15", "3 of 15", "12 of 15"],
        ["Assembler",
         "generic (ev_model sort + SB-quota cap + ≤3/nation + £100m)",
         "generic", "generic",
         "assemble_sb_hunter_squad — anchors 2× sure_shot_fwd + 1× influential_mid by ev_model, fills rest from sub-5% pool"],
        ["Captain rule", "argmax(ev_model × 2 + Σ ev_others over XI)", "same", "same", "same"],
        ["Budget mode", "Both (budget + unbudgeted)", "Both", "Both", "Both"],
        ["Tracking", "Each model runs as its own Challenge — round-by-round actuals tracked against fantasy_player_round_stats. Transfer rules per FIFA Fantasy spec: 2 free MD2/MD3, unlimited R32, 4 R16/QF, 5 SF, 6 Final, -3 per extra.",
         "same", "same", "same"],
    ]
    for j, h in enumerate(headers, start=1):
        s.cell(row=4, column=j, value=h)
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            s.cell(row=i, column=j, value=val)
    style_header_row(s, 4, len(headers))
    style_body(s, 5, 5 + len(rows) - 1, len(headers))

    next_r = 5 + len(rows) + 1
    s.cell(row=next_r, column=1, value="CURRENT ROUND SNAPSHOT").font = H2
    next_r += 1
    s.cell(row=next_r, column=1, value="Formation")
    for col, mid in enumerate(MIDS, start=2):
        s.cell(row=next_r, column=col, value=squad_by_id.get(mid, {}).get("formation", "—"))
    next_r += 1
    s.cell(row=next_r, column=1, value="Budget spent (£m)")
    for col, mid in enumerate(MIDS, start=2):
        s.cell(row=next_r, column=col, value=squad_by_id.get(mid, {}).get("budget_spent_m", 0))
    next_r += 1
    s.cell(row=next_r, column=1, value="SB-band picks (actual)")
    for col, mid in enumerate(MIDS, start=2):
        s.cell(row=next_r, column=col, value=squad_by_id.get(mid, {}).get("sb_band_count", 0))
    next_r += 1
    s.cell(row=next_r, column=1, value="Projected pts (incl. captain)")
    for col, mid in enumerate(MIDS, start=2):
        s.cell(row=next_r, column=col,
               value=round(squad_by_id.get(mid, {}).get("projected_pts_with_captain", 0), 1))
    next_r += 1
    s.cell(row=next_r, column=1, value="Captain")
    for col, mid in enumerate(MIDS, start=2):
        sq = squad_by_id.get(mid, {})
        cap_id = sq.get("captain_id")
        cap = next((p for p in sq.get("starting_xi", []) if p["fantasy_player_id"] == cap_id), None)
        s.cell(row=next_r, column=col, value=f"{cap['known_name']} ({cap['nation_id']})" if cap else "—")
    next_r += 1
    s.cell(row=next_r, column=1, value="12th Man")
    for col, mid in enumerate(MIDS, start=2):
        sq = squad_by_id.get(mid, {})
        t = sq.get("twelfth_man")
        s.cell(row=next_r, column=col, value=f"{t['known_name']} ({t['nation_id']}) £{t['price']}" if t else "—")
    style_body(s, 5 + len(rows) + 1, next_r, len(headers))
    set_widths(s, {1: 32, 2: 38, 3: 38, 4: 38, 5: 38})

    # Squad embedding
    section_r = next_r + 2
    s.cell(row=section_r, column=1, value="STARTING XI (current snapshot)").font = H2
    section_r += 1
    sub_headers = ["Slot"]
    for mid in MIDS:
        sub_headers.append(squad_by_id.get(mid, {}).get("name", mid))
    for j, h in enumerate(sub_headers, start=1):
        s.cell(row=section_r, column=j, value=h)
    style_header_row(s, section_r, len(sub_headers))
    section_r += 1
    for slot in range(11):
        s.cell(row=section_r, column=1, value=f"#{slot+1}")
        for col, mid in enumerate(MIDS, start=2):
            sq = squad_by_id.get(mid, {})
            xi = sq.get("starting_xi", [])
            if slot < len(xi):
                p = xi[slot]
                mark = "C" if p.get("captain") else ("VC" if p.get("vice_captain") else "")
                diff = "★" if p.get("is_differential") else ""
                s.cell(row=section_r, column=col,
                       value=f"[{p['position']}] {diff}{p['known_name']} ({p['nation_id']}) £{p['price']:.1f} {p['percent_selected']:.1f}% {mark}".strip())
        section_r += 1
    style_body(s, section_r - 11, section_r - 1, len(sub_headers))


# ─── Sheet 6: Chip strategy ──────────────────────────────────────────────────


def sheet_chips(wb: Workbook) -> None:
    s = wb.create_sheet("Chip strategy")
    add_title(s, "5 fantasy chips — EV computation + per-strategy timing")
    headers = ["Chip", "Rules / EV computation", "S1 Balanced Hunter",
               "S2 Steady Banker", "S3 Differential Maximizer"]
    rows = [
        ["Wildcard",
         "Unlimited transfers in one round. Cannot use at R32 (already unlimited). "
         "EV = E[squad with unlimited transfers − squad with capped transfers] across remaining rounds; fire at max delta.",
         "R16", "Group MD3 if 5+ injuries; else hold", "QF (variance amplifier)"],
        ["12th Man",
         "Pick one player NOT in squad to score for the round. No budget / nation cap. "
         "Can't be captained / subbed / transferred. EV = argmax(ev) over players NOT in 15.",
         "R16", "R16", "R32"],
        ["Maximum Captaincy",
         "Captain auto-assigned RETROACTIVELY to top-scoring XI player. "
         "EV = E[max(points across 11)] − 2·E[your captain pick]. Higher when XI has multiple high-ceiling players.",
         "QF", "QF", "SF"],
        ["Qualification Booster",
         "+2 per starting-XI player whose nation advances to next KO round. R32 onwards. "
         "EV = 2 × E[count of XI players whose nation advances] from Polymarket KO market.",
         "QF", "—", "R16"],
        ["Mystery Booster",
         "Revealed at R32 lock — hot-patch on reveal.",
         "TBD", "TBD", "TBD"],
    ]
    for j, h in enumerate(headers, start=1):
        s.cell(row=4, column=j, value=h)
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            s.cell(row=i, column=j, value=val)
    style_header_row(s, 4, len(headers))
    style_body(s, 5, 5 + len(rows) - 1, len(headers))
    set_widths(s, {1: 22, 2: 60, 3: 22, 4: 30, 5: 25})


# ─── Sheet 7: Anomalies & edge cases ─────────────────────────────────────────


def sheet_anomalies(wb: Workbook) -> None:
    s = wb.create_sheet("Anomalies & edge cases")
    add_title(s, "Anomalies, edge cases, and data-quality rules",
              "Surfaced explicitly so the model handles them rather than producing silently-wrong rows.")

    headers = ["Category", "Rule", "Handling"]
    rows = [
        ["Captain bonus", "Captain ×2 actual points. Vice steps in if captain plays 0 min.",
         "Assembler picks captain = argmax(ev × 2 + Σ ev_others) over starting XI. Captain choice IS part of EV."],
        ["Penalty shootouts", "Don't count for fantasy points (regulation + ET only).",
         "KO matches that go to shootouts cap point eligibility at 120-min final whistle. To confirm against FIFA 2026 ruleset."],
        ["Yellow accumulation", "Default: 2 yellows in group = 1-match ban; reset after QF.",
         "E2 filter drops suspended players. Pending FIFA confirmation."],
        ["KO TBD fixtures", "home_nation_id IS NULL until bracket fills.",
         "Skip with log line, not NaN-cascade. KO recommender only runs after bracket fills."],
        ["Fantasy ↔ FIFA ID mismatch", "fantasy_players.fifa_player_id null for ~10 players.",
         "Log orphan, exclude from recommendation set."],
        ["Position class disagreement", "fantasy.position (GK/DEF/MID/FWD) may disagree with stg_players_view.real_position.",
         "Trust fantasy position for scoring (FIFA scores against it); log delta."],
        ["Per-90 noise floor", "Per-90 stats off <90 min sample = noise.",
         "Mark low_confidence=true; UI dims the pick."],
        ["Goalscorer market staleness", "Polymarket last_update >24h old AND fixture <48h away → market stale.",
         "Demote A5 signal; fall back to fifa_wc_XG."],
        ["Debutant fallback", "fifa_wc_appearances=0 has no WC stats.",
         "Cascade to club_senior_* with 0.7 confidence haircut."],
        ["D1 inverse-ownership", "EDA §5: ownership correlates POSITIVELY with points (r=0.47).",
         "Differential reframed — value lives ONLY on the +2 SB event, gated at <5%, multiplied by sb_total track record."],
        ["Form autocorrelation", "fantasy.form is a derived rolling avg of recent points. Treat as lag-1 prior, not independent.",
         "Use recent5_fotmob_rating (r=0.46, independent) as the form-truth signal."],
        ["Squad assembly fallback", "If SB-band pool exhausted for some position quota, can't hit SB target.",
         "Track sb_gap; report rather than fail silently. Tolerance ±1 already built in."],
        ["Pre-match Polymarket snapshots (D-5)", "Closed-market calibration is post-event — useless for calibration.",
         "Notebook 13 cell 11 now appends to wc26_match_polymarket_markets_history.parquet with snapshot_ts."],
        ["Weather sparse sample", "Cluster 3 (wet) has n=2 matches in MD1+MD2.",
         "No modifier applied; revisit after MD3."],
        ["§I.3 heavy hitter", "upset_win_rate_career, knockout_overperformance need historical bracket data.",
         "Default to 1.0 multiplier (no-op). Backfill from hand-curated CSV in later phase."],
    ]
    for j, h in enumerate(headers, start=1):
        s.cell(row=4, column=j, value=h)
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            s.cell(row=i, column=j, value=val)
    style_header_row(s, 4, len(headers))
    style_body(s, 5, 5 + len(rows) - 1, len(headers))
    set_widths(s, {1: 26, 2: 55, 3: 65})


# ─── Sheet 8: Verification ───────────────────────────────────────────────────


def sheet_verification(wb: Workbook) -> None:
    s = wb.create_sheet("Verification")
    add_title(s, "Verification checklist + EDA correlation numbers",
              "What was eyeballed at each phase + the underlying signal evidence.")

    s["A4"] = "EDA SIGNAL — top factors by |corr| with actual fantasy points (MD1+MD2, n=1388)"
    s["A4"].font = H2
    eda = [
        ["Rank", "Factor", "DEF", "FWD", "GK", "MID", "max|corr|"],
        [1, "avg_points", 0.75, 0.82, 0.68, 0.83, 0.83],
        [2, "B8 goals_per_app", 0.31, 0.88, "—", 0.75, 0.88],
        [3, "form (autocorr)", 0.71, 0.80, 0.67, 0.77, 0.80],
        [4, "fifa_wc_Goals", 0.23, 0.72, "—", 0.64, 0.72],
        [5, "sb_total", 0.61, 0.54, 0.41, 0.68, 0.68],
        [6, "wc_rating / fotmob_rating", 0.47, 0.68, 0.49, 0.59, 0.68],
        [7, "fifa_wc_AttemptAtGoalOnTarget", 0.20, 0.59, "—", 0.44, 0.66],
        [8, "fifa_wc_XG", 0.17, 0.63, "—", 0.36, 0.63],
        [9, "B14 power_atk_score", 0.25, 0.63, "—", 0.63, 0.63],
        [10, "recent5_fotmob_rating (independent form)", 0.34, 0.46, 0.44, 0.39, 0.46],
        [11, "fifa_wc_CleanSheets", 0.38, 0.05, 0.44, 0.05, 0.44],
        [12, "fifa_wc_Assists", 0.21, 0.35, "—", 0.42, 0.42],
        [13, "Nation strength composite (A14)", "—", "—", "—", "—", 0.71],
        [14, "D1 percent_selected (POSITIVE)", 0.19, 0.47, 0.27, 0.20, 0.47],
        [15, "365 doubleChance trend ≥0.9 hit rate", "—", "—", "—", "—", "95%"],
    ]
    for j, h in enumerate(eda[0], start=1):
        s.cell(row=5, column=j, value=h)
    for i, row in enumerate(eda[1:], start=6):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 5, 7)
    style_body(s, 6, 5 + len(eda) - 1, 7)

    # Phase checklist
    next_r = 5 + len(eda) + 2
    s.cell(row=next_r, column=1, value="PHASE CHECKLIST").font = H2
    next_r += 1
    checklist = [
        ["Phase A — Foundation", "✓ Done",
         "Round metadata in PWA. wc_trophies.csv. wc26_stg_team_match_metrics (81 team-matches × 65 cols). lib/scores365.py + notebook 18 (459 trends, 62% baseline hit-rate)."],
        ["Phase B′ — EDA", "✓ Done",
         "17a_eda_factor_signal.py (820 lines). 9 analyses + extension v2 with §10 broad sweep (158 numeric cols), §11 prospective archetypes, §2b grass/roof overlay."],
        ["Phase C — Recommender", "✓ Done",
         "17_fantasy_recommender.py emits per-(player, fixture) rows in 3 modes + archetype enrichment + reason chips. 1248 rows for MD3."],
        ["Phase C.5 — Archetype v2", "✓ Done",
         "Retrospective k=6 (silhouette 0.22, MD1+MD2 top-20%). Prospective k=8 (silhouette 0.54, full pool). Messi/Haaland/Kane cluster as ELITE_FWD_GOALS_POPULAR."],
        ["Phase D — Strategy squads", "✓ Done",
         "3 squads: Balanced Hunter (10/15 diff, proj 67.9), Steady Banker (6/15, 78.2), Differential Max (12/15, 56.1). Hard SB-cap enforcement."],
        ["Phase D — Position Suggestor", "✓ Done",
         "Top 15 ranked overall + Look out for: 5 DEF + 5 MID + 5 FWD + 2 GK = 17 picks. Value score = ev/√(%sel+1) × (1 + sb_lift)."],
        ["Phase D-5 — Polymarket history append", "✓ Done",
         "Notebook 13 cell 11 patched. wc26_match_polymarket_markets_history.parquet captures snapshots for future calibration."],
        ["Phase E — PWA render", "Pending",
         "Loaders + components for K's 2 cents sub-tab. Football-pitch StrategySquad view. %selected slider that dims (not drops) above threshold."],
        ["Phase F — Round tracking", "Pending (post-MD3)",
         "Notebook 17 joins fantasy_player_round_stats; emits per (strategy, round) actuals."],
        ["Phase G — Excel deliverable", "✓ This file",
         "8 sheets covering catalog, lineage, scoring, strategies, chips, anomalies, verification."],
    ]
    s.cell(row=next_r, column=1, value="Phase")
    s.cell(row=next_r, column=2, value="Status")
    s.cell(row=next_r, column=3, value="Notes")
    style_header_row(s, next_r, 3)
    next_r += 1
    for row in checklist:
        for j, v in enumerate(row, start=1):
            s.cell(row=next_r, column=j, value=v)
        next_r += 1
    style_body(s, next_r - len(checklist), next_r - 1, 3)
    set_widths(s, {1: 32, 2: 16, 3: 70})


# ─── Sheet 9: Round Lock (determinism + freeze cadence) ────────────────────


def sheet_round_lock(wb: Workbook) -> None:
    s = wb.create_sheet("Round Lock")
    add_title(s, "Round-lock snapshot — deterministic re-runs",
              "Freezes every stat input at post-R(target-1) state. Re-runs for the same target round are byte-stable except for live %selected.")

    s["A4"] = "MECHANIC"
    s["A4"].font = H2
    mechanic = [
        ["#", "Step", "Detail"],
        [1, "Determine target round",
         "pick_target_round() reads LIVE fantasy_rounds.parquet. Picks the round in [start_date, end_date] (active by date), then next scheduled WITH fixtures, then last completed. R32 (R4) only goes live once fantasy_round_matches.parquet has 24 R4 entries — i.e. once the group stage finishes and the bracket fixtures lock."],
        [2, "Compute lock_round",
         "lock_round = max(0, target - 1). For target=R3 → lock=R2; for target=R4 (R32) → lock=R3; etc. Stats up to AND INCLUDING lock_round are used; everything beyond is filtered out."],
        [3, "Validate / create",
         "Snapshot dir: data/processed/locked/post_round_{lock_round:02d}/. If it exists AND has every parquet lib/recommender.py reads → reuse. If anything's missing (e.g. a new lib dependency landed) → wipe + re-create. Env RELOCK=1 forces a re-create."],
        [4, "Bulk-copy",
         "Every *.parquet from live PROC → lock_dir. Defensive default so new warehouse parquets don't break the next CI run."],
        [5, "Filter per-round source",
         "fantasy_player_round_stats.parquet → keep rows with round_id ≤ lock_round."],
        [6, "Filter per-match sources",
         "wc26_player_match_stats_wide.parquet, wc26_player_match_powerrank.parquet, wc26_stg_team_match_metrics.parquet — keep rows whose match_number is in the round_id ≤ lock_round window (mapped via fantasy_round_matches → squads → stg_matches)."],
        [7, "Rebuild aggregates from filtered sources",
         "wc26_stg_fantasy_player_totals (appearances / scouting_bonus / total_points / …), wc26_stg_player_powerrank (avg attacking/defensive/creativity/defending_goal_score), wc26_stg_players_view (all 55 fifa_wc_* SUM/AVG/MAX cols + fotmob_wc_appearances override using fifa_wc_n_matches with TimePlayed>0 filter — fixes the bench-row inflation bug)."],
        [8, "Rebind PROC",
         "lib.recommender.PROC = lock_dir for the scoring + archetype + fixture-profile + squad-assembly calls. Restored to live PROC before history-write so artefacts land in data/processed/ as usual."],
        [9, "Manifest",
         "_lock_manifest.json stamps target_round, lock_round, created_utc, locked_match_count, and which parquets were filtered vs rebuilt. Inspect this to debug stale locks."],
    ]
    for i, row in enumerate(mechanic, start=5):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 5, 3)
    style_body(s, 6, 5 + len(mechanic) - 1, 3)
    set_widths(s, {1: 5, 2: 28, 3: 90})

    # Current lock state from manifest if available
    next_r = 5 + len(mechanic) + 2
    s.cell(row=next_r, column=1, value="CURRENT LOCK STATE").font = H2
    next_r += 1
    manifest_path = PROC / "locked"
    if manifest_path.exists():
        locks = sorted(manifest_path.glob("post_round_*/_lock_manifest.json"))
        if locks:
            for mf_path in locks:
                try:
                    mf = json.loads(mf_path.read_text())
                except Exception:
                    continue
                s.cell(row=next_r, column=1, value=mf_path.parent.name)
                s.cell(row=next_r, column=2, value=f"target=R{mf.get('target_round')} lock=R{mf.get('lock_round')}")
                s.cell(row=next_r, column=3,
                       value=f"created {mf.get('created_utc','?')[:19]}Z · {mf.get('locked_match_count','?')} matches · {mf.get('total_parquets_in_lock','?')} parquets")
                next_r += 1
        else:
            s.cell(row=next_r, column=1, value="(no _lock_manifest.json found — first cron tick will create one)")
            next_r += 1
    else:
        s.cell(row=next_r, column=1, value="(locked/ directory not present — will be created on first run)")
        next_r += 1
    style_body(s, 5 + len(mechanic) + 2, next_r - 1, 3)

    # Confirmation table — answers the typical "is R3 really R2-only" question
    next_r += 1
    s.cell(row=next_r, column=1, value="DETERMINISM AUDIT").font = H2
    next_r += 1
    audit = [
        ["Question", "Answer"],
        ["R3 suggestions only use up-to-R2 stats?",
         "YES. fantasy_player_round_stats filtered to round_id ≤ 2; wide stats filtered to match_number ≤ 48; stg_fantasy_player_totals + stg_player_powerrank + stg_players_view fifa_wc_* rebuilt from those filtered sources. Cumulative tournament aggregates can no longer leak partial R3 data."],
        ["R32 (R4) will pick up MD3 stats automatically?",
         "YES. Once fantasy_round_matches has R4 fixtures (which lands when the group stage finishes), pick_target_round() flips to R4 and a fresh lock dir post_round_03/ is created including all R3 stats. No manual trigger."],
        ["Will re-runs of R3 change suggestions?",
         "NO for stats — locked. Only live %selected (refresh_live_percent_selected) shifts between runs, which intentionally moves the differential / SB-eligibility math by tiny amounts."],
        ["What about FotMob-only WC stats not in fifa_wc_*?",
         "CLOSED. Two-track rebuild inside the lock: (1) RECOVERABLE counters (goals, assists, minutes_played, yellows, reds, fotmob_rating) re-aggregated from wc26_player_recent_matches_fotmob filtered to WC2026 matches with match_date_utc ≤ lock window end. (2) NON-RECOVERABLE counters (chances_created, big_chances_created, dribbles, duels_won, touches, touches_opp_box, defensive_contributions, tackles, fouls_committed, xg_against_on_pitch) scaled by (lock_apps / live_apps) per player — assumes uniform per-match production, which is the standard unbiased estimator. Rates (duels_won_pct, successful_dribbles_pct) kept as-is — distribution-invariant under uniformity. All 6 fotmob_wc_* cols actually consumed by lib/recommender.py are now genuinely round-bounded."],
        ["Is the live %selected leak an issue for determinism?",
         "Intentional. The SB-eligibility gate uses live ownership so picks reflect current market. Stats are frozen; ownership floats."],
    ]
    for i, row in enumerate(audit):
        s.cell(row=next_r, column=1, value=row[0])
        s.cell(row=next_r, column=2, value=row[1])
        if i == 0:
            style_header_row(s, next_r, 2)
        next_r += 1
    style_body(s, next_r - len(audit) + 1, next_r - 1, 2)


# ─── Sheet 10: Archetypes (clusters from mine_archetypes_v2) ───────────────


def sheet_archetypes(wb: Workbook) -> None:
    s = wb.create_sheet("Archetypes")
    add_title(s, "Player archetype clusters",
              "K-means clustering in FIFA-stat + scoring-channel feature space. "
              "Two passes: retrospective (MD1+MD2 top-20% scorers) and prospective (full 1488 pool). "
              "Best-k chosen by silhouette across {6, 8, 10, 12}.")

    # Algorithm summary
    s["A4"] = "ALGORITHM"
    s["A4"].font = H2
    algo = [
        ["Step", "Detail"],
        ["Feature build",
         "Per-player aggregate from wc26_player_match_stats_wide (locked). p90 cols: AttemptAtGoal, AttemptAtGoalOnTarget, XG, Assists, Crosses, CrossesCompleted, PassesCompleted, Tackles, DefensivePressuresApplied, ForcedTurnovers, Corners, TotalDistance, GoalkeeperSaves. Plus wc_rating, FDH powerrank (attacking/defensive/creativity), price, %selected, scoring-channel composition (pts_from_goals/assists/cs/saves/sb/bonus pct)."],
        ["Drop nulls",
         "Any feature column with >40% null is dropped (avoids over-imputation in early-tournament samples). Remaining nulls filled with column median."],
        ["Scale + cluster",
         "StandardScaler + KMeans (n_init=10, random_state=42) at k ∈ {6,8,10,12}. Pick k with highest silhouette_score(Xs, labels)."],
        ["Name each cluster",
         "{TIER}_{POSITION}_{TOP_CHANNEL}_{TOP_STAT}_{OWN_BAND}. "
         "Tier from mean price (ELITE ≥9, MID_TIER ≥6, else BUDGET). "
         "Position = modal position. "
         "Top channel from scoring-channel composition (pts_from_goals / assists / cs / saves / sb / bonus). "
         "Top stat from highest mean p90 feature. "
         "Own band: DIFF if median %sel < 5 else POPULAR."],
        ["Exemplars",
         "Top-3 players per cluster by total fantasy points so far (or 0 if no R≤lock data yet)."],
        ["Attached to scoring frame",
         "attach_archetypes(scored, retro, prospective) writes peer_archetype_retrospective + peer_archetype_prospective + similarity + top-3 examples onto every recommendation row. PWA reads these as reason chips."],
    ]
    for i, row in enumerate(algo, start=5):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 5, 2)
    style_body(s, 6, 5 + len(algo) - 1, 2)
    set_widths(s, {1: 22, 2: 95})

    next_r = 5 + len(algo) + 2

    # Live cluster definitions
    for kind, fname in (("Retrospective (MD1+MD2 top-20%)", "archetypes_retrospective_v2.json"),
                        ("Prospective (full pool, pre-tournament profile)", "archetypes_prospective_v2.json")):
        s.cell(row=next_r, column=1, value=kind.upper()).font = H2
        next_r += 1
        eda_path = ROOT / "data" / "eda" / fname
        if not eda_path.exists():
            s.cell(row=next_r, column=1, value=f"(file {fname} not yet emitted — run 17_fantasy_recommender.py)")
            next_r += 2
            continue
        try:
            data = json.loads(eda_path.read_text())
        except Exception as e:
            s.cell(row=next_r, column=1, value=f"(failed to parse {fname}: {e})")
            next_r += 2
            continue
        meta_line = (f"k={data.get('k')} · silhouette={data.get('silhouette', 0):.3f} · "
                     f"{len(data.get('archetypes', []))} clusters · "
                     f"{len(data.get('feature_cols', []))} features")
        s.cell(row=next_r, column=1, value=meta_line).font = BODY_BOLD
        next_r += 1
        # Cluster table
        hdr = ["Cluster", "Name", "N", "Mean pts", "Top exemplars"]
        for j, h in enumerate(hdr, start=1):
            s.cell(row=next_r, column=j, value=h)
        style_header_row(s, next_r, len(hdr))
        first = next_r + 1
        next_r += 1
        for a in data.get("archetypes", []):
            s.cell(row=next_r, column=1, value=a.get("cluster_id"))
            s.cell(row=next_r, column=2, value=a.get("name"))
            s.cell(row=next_r, column=3, value=a.get("n"))
            s.cell(row=next_r, column=4, value=round(a.get("mean_pts", 0), 1))
            ex = a.get("exemplars") or []
            ex_txt = " · ".join(
                f"{e.get('known_name','?')} ({e.get('nation_id','')}) {int(e.get('total_pts',0))}pts"
                for e in ex
            )
            s.cell(row=next_r, column=5, value=ex_txt or "(no exemplars yet)")
            next_r += 1
        style_body(s, first, next_r - 1, 5)
        next_r += 1
    set_widths(s, {1: 10, 2: 38, 3: 6, 4: 10, 5: 80})


# ─── Sheet 11: Decision Flow (end-to-end ensemble pipeline) ────────────────


def sheet_decision_flow(wb: Workbook) -> None:
    s = wb.create_sheet("Decision Flow")
    add_title(s, "Decision flow — inputs → factors → brackets → models → joint → outputs",
              "What gets read from where, what gets computed, how the 4 models combine into PWA outputs.")

    s["A4"] = "INPUTS (under round lock)"
    s["A4"].font = H2
    inputs = [
        ["Group", "Parquet (in lock_dir)", "Used for"],
        ["Per-round fantasy", "fantasy_player_round_stats.parquet (filtered)",
         "Source of cumulative fantasy totals (rebuilt into stg_fantasy_player_totals). Drives B4 FantasyMeta — form, avg_pts, total_pts, consistency, last_round_points, scouting_bonus, sb_track multiplier."],
        ["Per-match FIFA", "wc26_player_match_stats_wide.parquet (filtered)",
         "53 spec'd FIFA stats per (player, match) — Goals, Assists, AttemptAtGoal(OnTarget), XG, Tackles, GoalkeeperSaves, CleanSheets, Passes(Completed), BallProgressions, SwitchesOfPlay, ChancesCreated, TouchesOppBox, DefensivePressures, etc. Rebuilt into stg_players_view.fifa_wc_* cols (SUM/AVG/MAX). Drives B2 WCPerfRating (per-90 + Bayesian shrinkage, position-routed)."],
        ["Per-match FDH powerrank", "wc26_player_match_powerrank.parquet (filtered)",
         "Attacking / defensive / creativity / defending_the_goal scores per (player, match). Rebuilt into stg_player_powerrank avg_*. Feeds B2 (30% weight) for position-conditional power-rank percentile."],
        ["Identity + profile", "fantasy_players.parquet, wc26_stg_players_view.parquet (career/value/recent-form cols preserved live)",
         "Position, price, %selected, known_name, fifa_player_id↔fantasy_player_id mapping. Career rollups (club_senior_*, national_senior_*), market value (latest/peak), FotMob recent5/10/15 windows. Drives B1 PlayerOverall + B3 ExternalRatings."],
        ["FotMob WC tournament rollup", "wc26_stg_players_view.fotmob_wc_* (live snapshot, partial leak)",
         "FotMob's tournament-stats endpoint — chances_created, big_chances, dribbles, duels_won, touches_opp_box, defensive_contributions, xg_against_on_pitch. Feeds B2 (FotMob branch). NOT round-filterable — see Round Lock sheet audit."],
        ["Fixtures + schedule", "fantasy_round_matches.parquet, wc26_stg_matches.parquet (as-is)",
         "Round 24-match window, home/away nation_id, kickoff_utc, stage, venue. Drives fixture cross-join in scoring."],
        ["Markets", "wc26_match_polymarket_markets.parquet, wc26_polymarket_match_volume.parquet (as-is)",
         "Per-fixture Polymarket implied probabilities: moneyline (p_home_win/p_away_win/p_draw), over/under totals (0.5-9.5), BTTS, per-side scoring. Feeds goals_index + cs_index + p_btts in fixture profile."],
        ["Trends + weather", "wc26_match_trends_365.parquet, wc26_match_weather.parquet (as-is)",
         "Scores365 historic-trend rankings per fixture; weather cluster + temperature + humidity + roof_type + surface. Feeds B5 FixtureMultiplier modifiers (weather_drag, trend_top_confidence)."],
        ["Nations identity", "wc26_stg_nations.parquet, fantasy_squads.parquet (as-is)",
         "Confederation, FIFA rank, pot, squad valuation. Drives nation_strength_composite in B1."],
        ["Team match metrics", "wc26_stg_team_match_metrics.parquet (filtered)",
         "Per-team WC2026 match metrics. Used by team-strength derivations."],
        ["Live %selected (NOT locked)", "refresh_live_percent_selected() — vercel /api/fifa-fantasy edge proxy",
         "Tick-fresh ownership for every fantasy_player_id. Overrides snapshot %selected before scoring. Drives differential / SB-eligibility math. Intentionally floats between re-runs of the same lock."],
    ]
    for i, row in enumerate(inputs, start=5):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, 5, 3)
    style_body(s, 6, 5 + len(inputs) - 1, 3)
    set_widths(s, {1: 26, 2: 42, 3: 78})

    next_r = 5 + len(inputs) + 2
    s.cell(row=next_r, column=1, value="ENSEMBLE FLOW").font = H2
    next_r += 1
    pipeline = [
        ["Step", "Function", "Output"],
        ["1. Archetype mining (shared)",
         "mine_archetypes_v2('retrospective') + mine_archetypes_v2('prospective')",
         "data/eda/archetypes_*_v2.json — cluster definitions + player_clusters map. Both runs read the LOCKED parquets."],
        ["2. Fixture profile (per round)",
         "assemble_fixture_profile(target_round)",
         "24 fixture rows: goals_index, cs_index_home/away, p_home_win/away_win/draw, p_btts, nation_strength_delta, fixture_shape ∈ {consensus_lopsided, consensus_tight, market_overconfident, composite_overconfident}, weather_cluster, trend_top_*."],
        ["3a. Score under each of 4 models",
         "score_for_model(target, fx, model_id) → score_players_brackets() under that model's weights",
         "Per-player rows with b1_overall, b2_wc_perf, b3_external, b4_fantasy, b5_fixture_mult, bracket_sum, ev_bracket, ev_model (post-boosts applied for M2/M3/M4)."],
        ["3b. Filters + anti-pick tag",
         "apply_filters() then tag_anti_picks()",
         "Drops is_active=False + injured. Tags 2nd/3rd ranked players on the opponent side of consensus_tight fixtures as anti_picks (D-3)."],
        ["3c. Reason chips",
         "tag_chips()",
         "HEDGE / DIFFERENTIAL / SB_TRACK_xN / SB_LIKELY / CEILING_HOT / FAVORED_FIXTURE / TREND_<CAT>."],
        ["3d. Attach archetypes",
         "attach_archetypes(scored, retro, prospective)",
         "archetype_retrospective + sim + 3 examples; archetype_prospective + sim + 3 examples."],
        ["4. Joint output",
         "build_joint_picks(model_outputs, top_n=30)",
         "consensus (in ≥2 models' top-30, sorted by max_ev), surprises (in exactly 1 model's top-30), per_position_top (15 FWD / 15 MID / 15 DEF / 5 GK by max ev_model across ALL models — pulls from FULL scored frame, not the top-30 union)."],
        ["5. Per-model squads",
         "assemble_strategy_squad (M1/M2/M3) or assemble_sb_hunter_squad (M4)",
         "One challenge squad per model: 2/5/5/3 quota, ≤3/nation, £100m budget, captain + vice + 12th-man. Also emits an unbudgeted variant for reference."],
        ["6. Position suggestor (legacy)",
         "build_position_suggestor(banker)",
         "wc26_fantasy_position_suggestor.json — kept for back-compat."],
        ["7. Round tracking",
         "build_round_tracking(squads_by_round)",
         "Closed-round actuals joined against fantasy_player_round_stats; running totals per model. Auto-pulls prior round snapshots from data/processed/history/round_NN/."],
        ["8. PWA emit",
         "_emit_pwa_json.py (downstream) + parquet→json mirroring inside 17",
         "Writes data/processed/json/*.json which the PWA fetches at /data/."],
    ]
    for i, row in enumerate(pipeline, start=next_r):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, next_r, 3)
    style_body(s, next_r + 1, next_r + len(pipeline) - 1, 3)
    next_r += len(pipeline) + 1

    next_r += 1
    s.cell(row=next_r, column=1, value="PWA RENDERING (consumer side)").font = H2
    next_r += 1
    pwa = [
        ["Section", "Detail"],
        ["Suggested Picks → By Position",
         "joint.per_position_top — 15 FWD / 15 MID / 15 DEF / 5 GK. Position chip filters to one position."],
        ["Suggested Picks → High Confidence",
         "joint.consensus — players in ≥2 model top-30 lists. Tile shows model attribution chips (M1/M2/M3/M4 short labels, no EV numbers — feedback removed the per-model EV)."],
        ["Suggested Picks → Per-Model Surprises",
         "joint.surprises grouped by model_id. Tile design same as consensus."],
        ["Suggested Picks → SB Eligible",
         "Deduped pool from joint.per_position_top + filter to live %sel < 5%. Honours position chips."],
        ["Fantasy Challenge",
         "4 strategy_squads (one per model). Tap to drill in: FIFA-style round-picker tabs across the top, total pts huge, captain marked, per-tile live/finished badge."],
    ]
    for i, row in enumerate(pwa, start=next_r):
        for j, v in enumerate(row, start=1):
            s.cell(row=i, column=j, value=v)
    style_header_row(s, next_r, 2)
    style_body(s, next_r + 1, next_r + len(pwa) - 1, 2)


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_factor_catalog(wb)
    sheet_data_lineage(wb)
    sheet_scoring(wb)
    sheet_strategies(wb)
    sheet_chips(wb)
    sheet_anomalies(wb)
    sheet_verification(wb)
    sheet_round_lock(wb)
    sheet_archetypes(wb)
    sheet_decision_flow(wb)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
