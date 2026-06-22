"""Walk every wc26/fantasy/referee parquet under data/processed/ and emit
WC26_DATA_DICTIONARY.xlsx — one sheet per table + an Overview sheet — with
column name, dtype, sample, null count, hand-written description."""
import sys, io as sio
sys.stdout = sio.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
PROCESSED = ROOT / "data" / "processed"
OUT = ROOT / "WC26_DATA_DICTIONARY.xlsx"

# ── Per-table notes (1-line description shown on the Overview + atop each sheet)
TABLE_NOTES: dict[str, dict] = {
    "wc26_nations":          {"notebook": "01", "refresh": "frozen",
        "desc": "Canonical 48-nation table for WC26. Carries the cross-source IDs (ESPN, FotMob, TM) and the alias union for nation-name resolution."},
    "wc26_stadiums":         {"notebook": "02", "refresh": "frozen",
        "desc": "16 host venues with capacity, roof, surface, altitude, lat/lng, IANA timezone, weather grid key, ESPN venue display string."},
    "wc26_matches":          {"notebook": "03", "refresh": "3h",
        "desc": "104 WC26 matches. ESPN status/score, FIFA IdMatch + IdIFES (fdh-api join), tactics, attendance, centre referee inline."},
    "wc26_player_match_stats":     {"notebook": "07", "refresh": "3h + event",
        "desc": "fdh-api per-match per-player stats, long format: one row per (match, player, stat_name). 53 curated stat keys (allowlist)."},
    "wc26_player_match_powerrank": {"notebook": "07", "refresh": "3h + event",
        "desc": "fdh-api per-match attacking/defensive/creativity ranks per player, plus GK defending/in-possession ranks."},
    "wc26_player_enrichment":            {"notebook": "08", "refresh": "daily",
        "desc": "One row per FIFA player with cross-source IDs (FotMob, TM), current market values (2 sources), contract end, preferred foot, current club, WC tournament line."},
    "wc26_player_market_value_history":  {"notebook": "08", "refresh": "daily",
        "desc": "FotMob scisports market-value series, long format: (player, date, value, lower/upper bound, team)."},
    "wc26_player_career_senior":         {"notebook": "08", "refresh": "daily",
        "desc": "FotMob senior club career — teamEntries (per-club spell totals) and seasonEntries (per-season totals)."},
    "wc26_player_career_national":       {"notebook": "08", "refresh": "daily",
        "desc": "FotMob national-team career, same shape as career_senior."},
    "wc26_player_fotmob_wc":             {"notebook": "11", "refresh": "3h",
        "desc": "FotMob WC tournament line per player. Carries apps, goals, assists, FotMob rating, plus 12 deep WC stats (chances_created, big_chances_created, dribbles, successful_dribbles_pct, duels_won/pct, touches, touches_opp_box, defensive_contributions, tackles, fouls_committed, xg_against_on_pitch) parsed from FotMob's firstSeasonStats — populated only when its match count matches the player's WC appearances."},
    "wc26_player_recent_matches_fotmob": {"notebook": "11", "refresh": "3h",
        "desc": "Per-player last-N matches from FotMob (any competition). Recent form features."},
    "wc26_match_weather":   {"notebook": "12", "refresh": "3h",
        "desc": "Per-match weather: temp, apparent temp, humidity, dew point, precipitation, wind, cloud cover, WMO code. WBGT proxy derived."},
    "wc26_polymarket_match_volume":   {"notebook": "13", "refresh": "3h",
        "desc": "One row per match. Polymarket event_ids (base + more-markets) clubbed, status flags, volume_moneyline (= moneyline + draw) vs volume_other (= over/under + spread + goalscorer)."},
    "wc26_match_polymarket_markets":  {"notebook": "13", "refresh": "3h",
        "desc": "One row per (event × market). 3,087 rows: moneyline (140) + draw (70) + over/under (2,053) + spread (608) + goalscorer (216). Each row carries last_trade_price (implied prob), volume, liquidity, spread, best_bid/ask."},
    "wc26_player_match_stats_wide": {"notebook": "07", "refresh": "3h + event",
        "desc": "Pivot of wc26_player_match_stats: one row per (match × player) × 116 stat columns. Easier to slice in Excel/SQL than the long form."},
    "referee_master":       {"notebook": "04", "refresh": "daily",
        "desc": "50 WC26 referees: name, country, confederation, FootyMetrics profile URL. wc26_nominated=true for all."},
    "referee_profile":      {"notebook": "04", "refresh": "daily",
        "desc": "Long: one row per (referee, source, window). Windows: career / last_10 / last_25 of FootyMetrics-tracked fixtures."},
    "referee_assignments":  {"notebook": "05", "refresh": "3h + event",
        "desc": "One row per (match, role, referee_id). Two sources: FIFA Officials primary + FootyMetrics upcoming-fixtures verification."},
    "fantasy_rounds":               {"notebook": "10", "refresh": "3h",
        "desc": "8 fantasy gameweeks with status + start/end UTC + per-round match count."},
    "fantasy_round_matches":        {"notebook": "10", "refresh": "3h",
        "desc": "Per-round fixture list, includes venue + live period + score."},
    "fantasy_squads":               {"notebook": "10", "refresh": "3h",
        "desc": "48 fantasy squads: fantasy_squad_id ↔ nation abbr + group + elimination flag."},
    "fantasy_players":              {"notebook": "10", "refresh": "3h",
        "desc": "1488 fantasy players. fifa_player_id direct join. Carries percent_selected (Scouting Premium input), price, form, total points."},
    "fantasy_player_round_stats":   {"notebook": "10", "refresh": "3h + event",
        "desc": "Per (fantasy_player_id, round_id): raw fantasy stat keys (GS/AS/MP/YC/RC/CS/GC/ST/T/CC/SB/…) + fantasy points scored."},
    "ref_id_bridge":                {"notebook": "04", "refresh": "daily",
        "desc": "Bridge from numeric FIFA OfficialId (fifa_referee_id, in wc26_matches) → slug referee_id (in referee_master). Built via surname+iso3 match; manual overrides in data/overrides/ref_id_overrides.csv applied last."},
    "wc26_stg_nations":             {"notebook": "14", "refresh": "frozen",
        "desc": "Pass-through staging table over wc26_nations. Stable downstream contract — exists so the audit-app PWA can pin to a column set independent of upstream rename churn."},
    "wc26_stg_stadiums":            {"notebook": "14", "refresh": "3h",
        "desc": "Per-stadium staging: wc26_stadiums base + match counters (days/total/completed/attendance) from wc26_matches + weather min/max/mean from wc26_match_weather."},
    "wc26_stg_referee_profile":     {"notebook": "14", "refresh": "daily",
        "desc": "referee_profile base (one row per ref × window) left-joined with referee_master identity columns (name, country, confederation, fm_id, slug, fm_url)."},
    "wc26_stg_matches":             {"notebook": "15", "refresh": "3h",
        "desc": "Per-match staging: wc26_matches base + joins of stadium fields, weather, home/away nation fields (prefixed home_/away_), referee fields via ref_id_bridge → referee_master, and Polymarket per-match volume."},
    "wc26_stg_players":             {"notebook": "16", "refresh": "3h",
        "desc": "Per-player staging (~180 cols). Base = wc26_player_enrichment; left-joined with the four summary parquets (career club + national + market value + FotMob WC); plus per-player WC aggregate from wc26_player_match_stats_wide (fifa_wc_* prefix); plus four recent-form windows (5/10/15/20) from wc26_player_recent_matches_fotmob."},
    "wc26_stg_fantasy_player_totals": {"notebook": "14", "refresh": "3h",
        "desc": "One row per fantasy_player_id. Pure groupby aggregation of fantasy_player_round_stats — tournament-to-date counters: appearances, minutes_played, starting_xi, total_points, total_goals_scored, total_assists, clean_sheets, saves, tackles, chances_created, shots_on_target, scouting_bonus, yellow_cards. No joins."},
    "wc26_stg_players_view":         {"notebook": "16", "refresh": "3h",
        "desc": "Slim curated view over wc26_stg_players (~115 hand-picked cols, dropping youth career + obscure FIFA-WC stats + recent20 window). Adds 13 derived completion/ratio percentages — reception breakdown, pass/ball-progression/switches/cross/distributions/line-breaks completion, %distance walking, %distance high-speed sprinting — plus fifa_wc_TotalCards = yellow + red. Consumer surface for the PWA + EV scorer."},
    "wc26_stg_player_powerrank":     {"notebook": "14", "refresh": "3h",
        "desc": "One row per (fifa_player_id, fifa_team_id). Pure groupby aggregation of wc26_player_match_powerrank — per-match-average FDH power-ranking scores across the player's WC matches: avg_attacking_score, avg_defensive_score, avg_creativity_score, avg_defending_the_goal_score (GK only — NaN for outfielders). Plus n_matches_ranked + player_kind context. No joins."},
}

# ── Per-column descriptions. Anything not listed gets an empty cell — easy
# to fill in later from your domain knowledge.
COL_DESC: dict[tuple[str, str], str] = {
    # ─── wc26_nations
    ("wc26_nations", "nation_id"):            "Canonical 3-letter FIFA code. Primary key for the warehouse. FK target for nation_id everywhere.",
    ("wc26_nations", "seed_name"):            "Display name from teams.ts (the audit-app seed; matches the FIFA Final Draw spelling).",
    ("wc26_nations", "iso_alpha2"):           "ISO 3166-1 alpha-2 country code (e.g. mx, gb-eng). Used for flag icons.",
    ("wc26_nations", "confederation"):        "FIFA confederation: AFC / CAF / CONCACAF / CONMEBOL / OFC / UEFA.",
    ("wc26_nations", "group"):                "WC26 group letter A–L (12 groups × 4 teams).",
    ("wc26_nations", "pot"):                  "Draw pot 1–4 (host-seeded for MEX/CAN/USA).",
    ("wc26_nations", "fifa_rank"):            "FIFA Men's Ranking at the time the seed was authored (mid-2026).",
    ("wc26_nations", "squad_valuation_m_eur"):"Approx total squad market value in € millions (mid-2026), from audit-app seed.",
    ("wc26_nations", "is_host"):              "True for MEX / CAN / USA.",
    ("wc26_nations", "espn_team_id"):         "ESPN numeric team id (use for ESPN scoreboard / team-schedule joins).",
    ("wc26_nations", "espn_name"):            "ESPN's displayName for the team.",
    ("wc26_nations", "espn_abbreviation"):    "ESPN's 3-letter abbreviation (usually matches FIFA code, occasionally differs).",
    ("wc26_nations", "fotmob_team_id"):       "FotMob numeric team id (use for /api/data/teams?id= and /api/data/playerData filters).",
    ("wc26_nations", "fotmob_name"):          "FotMob's display name for the team.",
    ("wc26_nations", "fotmob_short_name"):    "FotMob's short label.",
    ("wc26_nations", "tm_team_id"):           "Transfermarkt /verein/{id} — the national-team-as-club id, not the country page id.",
    ("wc26_nations", "tm_name"):              "Transfermarkt's display name.",
    ("wc26_nations", "tm_code"):              "Transfermarkt's internal team code (e.g. MEX1, GB1).",
    ("wc26_nations", "tm_slug"):              "Transfermarkt URL slug. Used to compose /{slug}/startseite/verein/{tm_team_id}.",
    ("wc26_nations", "stars"):                "List of marquee players from the audit-app seed.",
    ("wc26_nations", "all_names"):            "Union of every alias / source-specific name seen for this nation. Drives nation_match.match_to_canonical().",

    # ─── wc26_stadiums
    ("wc26_stadiums", "stadium_id"):          "Slug primary key (e.g. azteca, sofi). FK target for stadium_id in matches.",
    ("wc26_stadiums", "name"):                "Stadium display name.",
    ("wc26_stadiums", "city"):                "City (US uses metro label, e.g. 'New York / New Jersey').",
    ("wc26_stadiums", "state_or_region"):     "State (USA) or province (CAN) or region (MEX).",
    ("wc26_stadiums", "country"):             "Host country code: USA / CAN / MEX.",
    ("wc26_stadiums", "capacity"):            "WC26 seating capacity (post-overlay).",
    ("wc26_stadiums", "roof_type"):           "open / retractable / fixed.",
    ("wc26_stadiums", "surface"):             "grass (permanent) / grass_overlay (turf base + grass installed for WC).",
    ("wc26_stadiums", "altitude_m"):          "Approximate altitude in metres (hand-curated from FIFA spec).",
    ("wc26_stadiums", "latitude"):            "Stadium centroid latitude (4 decimals — Open-Meteo grid resolution).",
    ("wc26_stadiums", "longitude"):           "Stadium centroid longitude.",
    ("wc26_stadiums", "timezone"):            "IANA timezone (e.g. America/Mexico_City). Drives kickoff_local derivation.",
    ("wc26_stadiums", "weather_grid_key"):    "lat,lng rounded to 1 decimal — deterministic key for the Open-Meteo grid cell.",
    ("wc26_stadiums", "open_meteo_elevation_m"):"Open-Meteo's reported elevation (sanity-check vs altitude_m).",
    ("wc26_stadiums", "altitude_delta_m"):    "open_meteo_elevation_m - altitude_m. Large deltas flag mis-typed altitudes.",
    ("wc26_stadiums", "match_key"):           "Prefix string used to match against ESPN/FIFA venue.fullName.",
    ("wc26_stadiums", "espn_venue_name"):     "ESPN's exact venue string (e.g. 'Estadio Banorte' for Azteca's sponsor name).",

    # ─── wc26_matches
    ("wc26_matches", "match_number"):         "Chronological match index 1..104 (kicked-off-first first).",
    ("wc26_matches", "espn_match_id"):        "ESPN's numeric event id. Primary external id for live state.",
    ("wc26_matches", "fifa_match_id"):        "FIFA api.fifa.com IdMatch (use with /calendar/matches).",
    ("wc26_matches", "fifa_id_ifes"):         "FIFA fdh-api IdIFES. **Required key for /v1/stats/match/{}/players.json and /v1/powerranking/match/{}.json.**",
    ("wc26_matches", "seed_match_id"):        "fixtures.ts integer id (group stage only — survives even if ESPN's id changes).",
    ("wc26_matches", "kickoff_utc"):          "Match kickoff in UTC. Source of truth for scheduling.",
    ("wc26_matches", "kickoff_local"):        "Kickoff in the stadium's IANA timezone — used for weather hour lookup.",
    ("wc26_matches", "date_et"):              "Date in America/New_York (legacy seed convention).",
    ("wc26_matches", "kickoff_et"):           "ET clock time HH:MM (seed-only; ESPN supersedes for live).",
    ("wc26_matches", "stage"):                "group_a..l / r32 / r16 / qf / sf / third_place / final.",
    ("wc26_matches", "espn_season_slug"):     "Raw ESPN season slug (group-stage / round-of-32 / ... / final).",
    ("wc26_matches", "home_nation_id"):       "FK → wc26_nations.nation_id. Null for knockout TBD slots until bracket fills.",
    ("wc26_matches", "away_nation_id"):       "FK → wc26_nations.nation_id.",
    ("wc26_matches", "fifa_home_team_id"):    "FIFA's IdTeam for home — required for /teams/{id}/squad calls (notebook 06).",
    ("wc26_matches", "fifa_away_team_id"):    "FIFA's IdTeam for away.",
    ("wc26_matches", "fifa_home_tactics"):    "Home formation string (e.g. '4-2-3-1'). Published post-lineup.",
    ("wc26_matches", "fifa_away_tactics"):    "Away formation.",
    ("wc26_matches", "stadium_id"):           "FK → wc26_stadiums.stadium_id.",
    ("wc26_matches", "fifa_stadium_id"):      "FIFA's IdStadium (parallel id; we use our slug-based stadium_id for joins).",
    ("wc26_matches", "espn_venue_name"):      "ESPN venue display (carried through for traceability).",
    ("wc26_matches", "venue_string"):         "Seed venue string from fixtures.ts.",
    ("wc26_matches", "status"):               "scheduled / live / finished.",
    ("wc26_matches", "home_score"):           "Home final score (or current live score).",
    ("wc26_matches", "away_score"):           "Away final score.",
    ("wc26_matches", "fifa_attendance"):      "FIFA-reported attendance (post-match).",
    ("wc26_matches", "fifa_referee_id"):      "FIFA OfficialId for the centre referee. Joins fantasy/FIFA referee tables.",
    ("wc26_matches", "fifa_referee_country"): "Referee nationality (FIFA 3-letter).",
    ("wc26_matches", "fifa_referee_name"):    "Referee display name from FIFA (CAPS surname convention).",
    ("wc26_matches", "espn_status_raw"):      "Raw ESPN status_type.name (for debugging mapper coverage).",
    ("wc26_matches", "espn_notes"):           "ESPN notes/headline (used as a fallback stage signal).",

    # ─── wc26_player_match_stats (long)
    ("wc26_player_match_stats", "fifa_match_id"):  "FK → wc26_matches.fifa_match_id (FIFA's IdMatch).",
    ("wc26_player_match_stats", "fifa_id_ifes"):   "FK → wc26_matches.fifa_id_ifes (fdh-api match id).",
    ("wc26_player_match_stats", "fifa_player_id"): "FK → wc26_players.",
    ("wc26_player_match_stats", "stat_name"):      "fdh-api stat key. ~116 distinct keys (Goals, Assists, BallProgressions, DistanceWalking, etc.).",
    ("wc26_player_match_stats", "value"):          "Numeric stat value (int or float). Distances are metres, durations are seconds, speeds are km/h.",

    # ─── wc26_player_match_powerrank
    ("wc26_player_match_powerrank", "fifa_match_id"): "FK → wc26_matches.fifa_match_id.",
    ("wc26_player_match_powerrank", "fifa_id_ifes"):  "FK → wc26_matches.fifa_id_ifes.",
    ("wc26_player_match_powerrank", "fifa_player_id"):"FK → wc26_players.",
    ("wc26_player_match_powerrank", "fifa_team_id"):  "Player's team (home or away).",
    ("wc26_player_match_powerrank", "player_kind"):   "outfieldPlayer / goalkeeper.",
    ("wc26_player_match_powerrank", "attacking_rank"):    "1-based rank vs all players in the match (1 = best attacker).",
    ("wc26_player_match_powerrank", "defensive_rank"):    "1-based attacking-vs-defensive symmetric.",
    ("wc26_player_match_powerrank", "creativity_rank"):   "1-based.",
    ("wc26_player_match_powerrank", "attacking_score"):   "Raw score (float). Higher = better.",
    ("wc26_player_match_powerrank", "defensive_score"):   "Raw score.",
    ("wc26_player_match_powerrank", "creativity_score"):  "Raw score.",
    ("wc26_player_match_powerrank", "attacking_rank_within_team"): "1-based rank within own team.",
    ("wc26_player_match_powerrank", "defensive_rank_within_team"): "1-based within team.",
    ("wc26_player_match_powerrank", "creativity_rank_within_team"):"1-based within team.",
    ("wc26_player_match_powerrank", "defending_the_goal_rank"):  "GK-specific. 1-based rank for goal-line defending.",
    ("wc26_player_match_powerrank", "defending_the_goal_score"): "GK-specific raw score.",
    ("wc26_player_match_powerrank", "in_possession_rank"):       "GK-specific. Build-up / distribution.",
    ("wc26_player_match_powerrank", "in_possession_score"):      "GK-specific raw score.",

    # ─── wc26_player_enrichment
    ("wc26_player_enrichment", "nation_id"):              "FK → wc26_nations.",
    ("wc26_player_enrichment", "fifa_player_id"):         "FK → wc26_players.",
    ("wc26_player_enrichment", "name"):                   "FIFA display name (carried for human readability when joining).",
    ("wc26_player_enrichment", "short_name"):             "FIFA short name.",
    ("wc26_player_enrichment", "birth_date"):             "Carried from wc26_players.",
    ("wc26_player_enrichment", "jersey_num"):             "Carried.",
    ("wc26_player_enrichment", "height_cm"):              "Carried.",
    ("wc26_player_enrichment", "weight_kg"):              "Carried.",
    ("wc26_player_enrichment", "position"):               "Carried.",
    ("wc26_player_enrichment", "real_position"):          "Carried.",
    ("wc26_player_enrichment", "real_position_side"):     "Carried.",
    ("wc26_player_enrichment", "preferred_foot"):         "From FotMob (FIFA doesn't publish this).",
    ("wc26_player_enrichment", "picture_url"):            "Carried.",
    ("wc26_player_enrichment", "fotmob_player_id"):       "FotMob numeric player id. Use for /api/data/playerData?id= and recentMatches.",
    ("wc26_player_enrichment", "fotmob_name"):            "FotMob display name (e.g. 'Heung-Min Son' vs FIFA 'SON Heungmin').",
    ("wc26_player_enrichment", "club_fotmob_id"):         "Current club's FotMob id (string).",
    ("wc26_player_enrichment", "club_name"):              "Current club name (FotMob).",
    ("wc26_player_enrichment", "transfer_value_eur_fotmob"):"FotMob's `transferValue` from the team page (scisports-modelled, EUR).",
    ("wc26_player_enrichment", "position_ids_desc"):      "Comma-separated FotMob position labels (e.g. 'ST,RW,CAM').",
    ("wc26_player_enrichment", "wc_rating"):              "FotMob WC tournament rating (0–10, this tournament).",
    ("wc26_player_enrichment", "wc_goals"):               "FotMob WC goals this tournament.",
    ("wc26_player_enrichment", "wc_assists"):             "FotMob WC assists.",
    ("wc26_player_enrichment", "wc_yellow_cards"):        "FotMob WC yellows.",
    ("wc26_player_enrichment", "wc_red_cards"):           "FotMob WC reds.",
    ("wc26_player_enrichment", "tm_player_id"):           "Transfermarkt numeric player id. Joins /spieler/{id}.",
    ("wc26_player_enrichment", "tm_slug"):                "TM URL slug.",
    ("wc26_player_enrichment", "club_tm_id"):             "Current club's TM id.",
    ("wc26_player_enrichment", "club_name_tm"):           "Current club name (TM).",
    ("wc26_player_enrichment", "market_value_eur_tm"):    "TM current market value in EUR (human-curated).",
    ("wc26_player_enrichment", "contract_end"):           "Contract end date from FotMob.",
    ("wc26_player_enrichment", "market_value_latest_eur_fotmob"): "FotMob's latest market value entry (scisports).",
    ("wc26_player_enrichment", "market_value_lower_eur_fotmob"):  "FotMob's confidence-interval lower bound.",
    ("wc26_player_enrichment", "market_value_upper_eur_fotmob"):  "FotMob's confidence-interval upper bound.",
    ("wc26_player_enrichment", "market_value_team_id_fotmob"):    "Team id associated with the latest FotMob value (current employer).",

    # ─── wc26_player_market_value_history
    ("wc26_player_market_value_history", "fifa_player_id"):  "FK → wc26_players.",
    ("wc26_player_market_value_history", "fotmob_player_id"):"FK → enrichment.",
    ("wc26_player_market_value_history", "date"):            "ISO date the valuation was made.",
    ("wc26_player_market_value_history", "value_eur"):       "Modelled valuation in EUR.",
    ("wc26_player_market_value_history", "lower_eur"):       "Lower bound.",
    ("wc26_player_market_value_history", "upper_eur"):       "Upper bound.",
    ("wc26_player_market_value_history", "team_id"):         "FotMob team id at the time.",
    ("wc26_player_market_value_history", "team_name"):       "Team name at the time.",
    ("wc26_player_market_value_history", "is_period_start"): "True when the valuation marks the start of a new club spell.",
    ("wc26_player_market_value_history", "source"):          "Almost always 'scisports'.",

    # ─── wc26_player_career_senior / national
    ("wc26_player_career_senior", "fifa_player_id"): "FK.",
    ("wc26_player_career_senior", "fotmob_player_id"):"FK.",
    ("wc26_player_career_senior", "kind"):           "'team' for per-club spell totals, 'season' for per-season rows.",
    ("wc26_player_career_senior", "fotmob_team_id"): "FotMob team id (only for kind='team').",
    ("wc26_player_career_senior", "team_name"):      "Club name.",
    ("wc26_player_career_senior", "transfer_type"):  "Transfer type when known (loan/permanent).",
    ("wc26_player_career_senior", "start_date"):     "Spell start date.",
    ("wc26_player_career_senior", "end_date"):       "Spell end (null = ongoing).",
    ("wc26_player_career_senior", "active"):         "True if currently at this club.",
    ("wc26_player_career_senior", "appearances"):    "Apps for the club / in the season.",
    ("wc26_player_career_senior", "goals"):          "Goals.",
    ("wc26_player_career_senior", "assists"):       "Assists.",
    ("wc26_player_career_senior", "has_uncertain_data"):"FotMob flag for data freshness.",
    ("wc26_player_career_senior", "season_name"):    "Only for kind='season' (e.g. '2024/2025').",
    ("wc26_player_career_senior", "rating"):         "FotMob season rating (only kind='season').",

    # ─── wc26_player_career_national — same schema
    ("wc26_player_career_national", "fifa_player_id"): "FK.",
    ("wc26_player_career_national", "fotmob_player_id"):"FK.",
    ("wc26_player_career_national", "kind"):           "team / season.",
    ("wc26_player_career_national", "team_name"):      "National team name.",
    ("wc26_player_career_national", "appearances"):    "Caps.",
    ("wc26_player_career_national", "goals"):          "International goals.",
    ("wc26_player_career_national", "assists"):       "Assists.",
    ("wc26_player_career_national", "season_name"):    "International calendar year for kind='season'.",

    # ─── wc26_player_fotmob_wc
    ("wc26_player_fotmob_wc", "fifa_player_id"):       "FK.",
    ("wc26_player_fotmob_wc", "fotmob_player_id"):     "FK.",
    ("wc26_player_fotmob_wc", "season_name"):          "Always '2026' for WC26.",
    ("wc26_player_fotmob_wc", "fotmob_tournament_id"): "FotMob's tournamentId (e.g. 24254 for WC26).",
    ("wc26_player_fotmob_wc", "appearances"):          "WC apps to date.",
    ("wc26_player_fotmob_wc", "goals"):                "WC goals.",
    ("wc26_player_fotmob_wc", "assists"):              "WC assists.",
    ("wc26_player_fotmob_wc", "fotmob_rating"):        "WC rating (0–10).",

    # ─── wc26_player_recent_matches_fotmob
    ("wc26_player_recent_matches_fotmob", "fifa_player_id"):    "FK.",
    ("wc26_player_recent_matches_fotmob", "fotmob_player_id"):  "FK.",
    ("wc26_player_recent_matches_fotmob", "match_date_utc"):    "Match kickoff UTC.",
    ("wc26_player_recent_matches_fotmob", "match_id_fotmob"):   "FotMob match id (links to /matches/{slug}#{id}).",
    ("wc26_player_recent_matches_fotmob", "team_id"):           "Player's team for this fixture.",
    ("wc26_player_recent_matches_fotmob", "team_name"):         "Team name.",
    ("wc26_player_recent_matches_fotmob", "opponent_team_id"):  "Opponent team id.",
    ("wc26_player_recent_matches_fotmob", "opponent_team_name"):"Opponent name.",
    ("wc26_player_recent_matches_fotmob", "is_home"):           "True if player's team was home.",
    ("wc26_player_recent_matches_fotmob", "league_id"):         "FotMob league id (77 = WC).",
    ("wc26_player_recent_matches_fotmob", "league_name"):       "League display name.",
    ("wc26_player_recent_matches_fotmob", "stage"):             "Stage label (knockout etc.).",
    ("wc26_player_recent_matches_fotmob", "home_score"):        "Final home score.",
    ("wc26_player_recent_matches_fotmob", "away_score"):        "Final away score.",
    ("wc26_player_recent_matches_fotmob", "minutes_played"):    "Player minutes.",
    ("wc26_player_recent_matches_fotmob", "goals"):             "Player goals in this match.",
    ("wc26_player_recent_matches_fotmob", "assists"):           "Player assists.",
    ("wc26_player_recent_matches_fotmob", "yellow_cards"):      "Player yellows.",
    ("wc26_player_recent_matches_fotmob", "red_cards"):         "Player reds.",
    ("wc26_player_recent_matches_fotmob", "fotmob_rating"):     "FotMob match rating (0–10).",
    ("wc26_player_recent_matches_fotmob", "is_top_rating"):     "True if highest rating across both teams.",
    ("wc26_player_recent_matches_fotmob", "player_of_the_match"):"True if POTM.",
    ("wc26_player_recent_matches_fotmob", "on_bench"):          "True if unused substitute.",

    # ─── wc26_match_weather
    ("wc26_match_weather", "espn_match_id"):        "FK → wc26_matches.",
    ("wc26_match_weather", "fifa_match_id"):        "Parallel FIFA id.",
    ("wc26_match_weather", "match_number"):         "1..104 chronological.",
    ("wc26_match_weather", "stadium_id"):           "FK → wc26_stadiums.",
    ("wc26_match_weather", "kickoff_utc"):          "Match kickoff UTC.",
    ("wc26_match_weather", "local_date"):           "Date in venue's local timezone.",
    ("wc26_match_weather", "local_hour"):           "Hour (0–23) in venue's local timezone — Open-Meteo lookup key.",
    ("wc26_match_weather", "source"):               "'archive' (observed, past matches) or 'forecast' (predicted, upcoming).",
    ("wc26_match_weather", "temperature_c"):        "Air temperature at 2m in °C.",
    ("wc26_match_weather", "apparent_temperature_c"):"Apparent (feels-like) temperature.",
    ("wc26_match_weather", "humidity_pct"):         "Relative humidity (0–100).",
    ("wc26_match_weather", "dew_point_c"):          "Dew point in °C.",
    ("wc26_match_weather", "precipitation_mm"):     "Hourly precipitation total (mm).",
    ("wc26_match_weather", "rain_mm"):              "Hourly rain (mm), subset of precipitation.",
    ("wc26_match_weather", "wind_speed_kmh"):       "Wind speed at 10m (km/h).",
    ("wc26_match_weather", "wind_direction_deg"):   "Wind direction (degrees, 0=N).",
    ("wc26_match_weather", "cloud_cover_pct"):      "Cloud cover percentage.",
    ("wc26_match_weather", "wmo_weather_code"):     "WMO weather code (see Open-Meteo docs for the enum).",
    ("wc26_match_weather", "wet_bulb_temp_c"):      "Stull (2011) wet-bulb temperature in °C, derived from temp + humidity.",
    ("wc26_match_weather", "wbgt_proxy_c"):         "WBGT proxy = 0.7·Tw + 0.2·Tappar + 0.1·T. >28 = heat-stress concern; >32 = severe.",

    # ─── wc26_team_officials
    ("wc26_team_officials", "nation_id"):     "FK → wc26_nations.",
    ("wc26_team_officials", "fifa_team_id"):  "FK → nations.",
    ("wc26_team_officials", "fifa_coach_id"): "FIFA IdCoach (primary key for the coach).",
    ("wc26_team_officials", "name"):          "Coach / official display name.",
    ("wc26_team_officials", "alias"):         "Alternate name when FIFA records it (e.g. nickname).",
    ("wc26_team_officials", "role"):          "Numeric role enum (FIFA's classification).",
    ("wc26_team_officials", "role_localized"):"Role in English (Head Coach / Assistant / etc.).",
    ("wc26_team_officials", "birth_date"):    "Coach DOB.",
    ("wc26_team_officials", "fifa_id_country"):"Coach nationality.",

    # ─── referee_master
    ("referee_master", "referee_id"):       "Slug-form key: lowercased-name-iso (e.g. 'szymon-marciniak-pl').",
    ("referee_master", "name"):             "Display name (FootyMetrics).",
    ("referee_master", "country"):          "FootyMetrics country.",
    ("referee_master", "confederation"):    "AFC / CAF / CONCACAF / CONMEBOL / OFC / UEFA.",
    ("referee_master", "flag_iso"):         "ISO 3166-1 alpha-2 country code for flag rendering.",
    ("referee_master", "nation_id"):        "FK → wc26_nations.nation_id when the ref is from a WC26 nation; null otherwise.",
    ("referee_master", "fm_id"):            "FootyMetrics numeric id (joins /referees/{fm_id}-{slug}).",
    ("referee_master", "slug"):             "FootyMetrics URL slug.",
    ("referee_master", "fm_url"):           "Full FootyMetrics profile URL.",
    ("referee_master", "countryApid"):      "FootyMetrics internal country id.",
    ("referee_master", "wc26_nominated"):   "Always True (this table is the WC26 panel).",

    # ─── referee_profile
    ("referee_profile", "referee_id"):           "FK → referee_master.",
    ("referee_profile", "source"):               "'footymetrics' currently (Soccerway integration is a follow-up).",
    ("referee_profile", "window"):               "career / last_10 / last_25.",
    ("referee_profile", "matches"):              "Match count in the window.",
    ("referee_profile", "yellow_pg"):            "Yellow cards per game.",
    ("referee_profile", "red_pg"):               "Red cards per game.",
    ("referee_profile", "penalty_pg"):           "Penalties awarded per game.",
    ("referee_profile", "fouls_pg"):             "Fouls called per game.",
    ("referee_profile", "added_time_fh_avg"):    "Average first-half added time (minutes). Career only.",
    ("referee_profile", "added_time_sh_avg"):    "Average second-half added time.",
    ("referee_profile", "booking_points_pg"):    "Booking points per game (yellow=10, red=25 per FA convention).",
    ("referee_profile", "total_yellows"):        "Window total yellows.",
    ("referee_profile", "total_reds"):           "Window total reds.",
    ("referee_profile", "total_penalties"):      "Window total penalties.",
    ("referee_profile", "total_fouls"):          "Window total fouls.",
    ("referee_profile", "fixtures_with_red"):    "Count of fixtures with ≥1 red.",
    ("referee_profile", "fixtures_with_penalty"):"Count of fixtures with ≥1 penalty.",
    ("referee_profile", "fixtures_no_cards"):    "Count of fixtures with zero cards.",
    ("referee_profile", "computed_at"):          "When this row was computed (UTC ISO).",

    # ─── referee_assignments
    ("referee_assignments", "match_id"):           "FK → wc26_matches.espn_match_id.",
    ("referee_assignments", "role"):               "referee / ar1 / ar2 / fourth / var / avar.",
    ("referee_assignments", "referee_id"):         "FK → referee_master.",
    ("referee_assignments", "source"):             "'fifa' (primary) or 'footymetrics' (verification overlay).",
    ("referee_assignments", "announced_at"):       "When the appointment published (null when unknown).",
    ("referee_assignments", "fm_fixture_id"):      "FootyMetrics fixture id when source=='footymetrics'.",
    ("referee_assignments", "fm_fixture_slug"):    "FM fixture slug for traceability.",
    ("referee_assignments", "home_nation_id"):     "FK.",
    ("referee_assignments", "away_nation_id"):     "FK.",
    ("referee_assignments", "match_number"):       "1..104.",
    ("referee_assignments", "kickoff_utc"):        "Match kickoff.",
    ("referee_assignments", "stage"):              "Group / knockout label.",
    ("referee_assignments", "fifa_official_id"):   "FIFA OfficialId (only when source=='fifa').",
    ("referee_assignments", "fifa_official_name"): "FIFA-published name.",

    # ─── fantasy_rounds
    ("fantasy_rounds", "round_id"):     "Fantasy gameweek id (1..8).",
    ("fantasy_rounds", "status"):       "upcoming / playing / complete.",
    ("fantasy_rounds", "start_date"):   "Round open date (kickoff window start).",
    ("fantasy_rounds", "end_date"):     "Round close date.",
    ("fantasy_rounds", "match_count"):  "Fixtures in this round.",

    # ─── fantasy_round_matches
    ("fantasy_round_matches", "round_id"):           "FK → fantasy_rounds.",
    ("fantasy_round_matches", "fantasy_match_id"):   "Fantasy's own match id.",
    ("fantasy_round_matches", "venue_id"):           "Fantasy venue id (1..16).",
    ("fantasy_round_matches", "venue_name"):         "Venue display.",
    ("fantasy_round_matches", "venue_city"):         "City.",
    ("fantasy_round_matches", "date"):               "Kickoff timestamp (ISO with tz).",
    ("fantasy_round_matches", "status"):             "scheduled / live / complete.",
    ("fantasy_round_matches", "period"):             "full_time / first_half / second_half / etc.",
    ("fantasy_round_matches", "minutes"):            "Live minutes.",
    ("fantasy_round_matches", "extra_minutes"):      "Stoppage time.",
    ("fantasy_round_matches", "is_suspended"):       "True for postponed fixtures.",
    ("fantasy_round_matches", "home_squad_id"):      "FK → fantasy_squads.",
    ("fantasy_round_matches", "home_squad_name"):    "Home team name.",
    ("fantasy_round_matches", "away_squad_id"):      "FK → fantasy_squads.",
    ("fantasy_round_matches", "away_squad_name"):    "Away team name.",
    ("fantasy_round_matches", "home_score"):         "Home score (final or live).",
    ("fantasy_round_matches", "away_score"):         "Away score.",

    # ─── fantasy_squads
    ("fantasy_squads", "fantasy_squad_id"): "Fantasy squad primary key (1..48).",
    ("fantasy_squads", "name"):             "Nation name.",
    ("fantasy_squads", "group"):            "Group letter a..l (lowercase in Fantasy).",
    ("fantasy_squads", "abbr"):             "3-letter abbr — joins wc26_nations.nation_id.",
    ("fantasy_squads", "is_eliminated"):    "True after knockout exit.",

    # ─── fantasy_players (the big one)
    ("fantasy_players", "fantasy_player_id"):  "Fantasy primary key. Required for fantasy_player_round_stats.fantasy_player_id.",
    ("fantasy_players", "fifa_player_id"):     "FK → wc26_players. Provided directly by Fantasy API — no fuzzy matching.",
    ("fantasy_players", "fantasy_squad_id"):   "FK → fantasy_squads.",
    ("fantasy_players", "first_name"):         "Given name.",
    ("fantasy_players", "last_name"):          "Family name.",
    ("fantasy_players", "known_name"):         "Mononym / nickname when used (e.g. 'Vinicius Jr').",
    ("fantasy_players", "position"):           "GK / DEF / MID / FWD (Fantasy's coarse position).",
    ("fantasy_players", "price"):              "Fantasy price (float, e.g. 11.5).",
    ("fantasy_players", "status"):             "playing / injured / suspended / etc.",
    ("fantasy_players", "match_status"):       "start / subbed / on / off / null (live).",
    ("fantasy_players", "percent_selected"):   "**Scouting Premium input.** Percentage of fantasy managers who have this player in their team.",
    ("fantasy_players", "rounds_selected_json"):"Per-round ownership history as JSON {round_id: percent}.",
    ("fantasy_players", "total_points"):       "Cumulative fantasy points season-to-date.",
    ("fantasy_players", "avg_points"):         "Average points per appearance.",
    ("fantasy_players", "form"):               "Fantasy's own form score.",
    ("fantasy_players", "last_round_points"):  "Points scored in the most-recent round.",
    ("fantasy_players", "round_points_json"):  "Per-round points as JSON {round_id: points}.",
    ("fantasy_players", "next_fixture_from_active_round"): "Fantasy match_id of the player's next fixture if the active round is in progress.",
    ("fantasy_players", "next_fixture_from_scheduled_round"):"Next fixture id from the next scheduled round.",
    ("fantasy_players", "one_to_watch"):       "True if FIFA's editorial flagged this player.",
    ("fantasy_players", "one_to_watch_text"):  "Editorial blurb (null when not flagged).",
    ("fantasy_players", "qualification_round_ids_json"):"Round ids the player qualified for, as JSON list.",

    # ─── fantasy_player_round_stats
    ("fantasy_player_round_stats", "fantasy_player_id"): "FK → fantasy_players. (Join to wc26 via fantasy_players.fifa_player_id.)",
    ("fantasy_player_round_stats", "round_id"):          "FK → fantasy_rounds.",
    ("fantasy_player_round_stats", "tournament_id"):     "Fantasy's per-team tournament id (one per (squad, round)).",
    ("fantasy_player_round_stats", "points"):            "**Fantasy points scored in this round** (the EV ground-truth).",
    ("fantasy_player_round_stats", "SXI"):               "Starting-XI flag (0/1).",
    ("fantasy_player_round_stats", "MP"):                "Minutes played.",
    ("fantasy_player_round_stats", "AS"):                "Assists.",
    ("fantasy_player_round_stats", "YC"):                "Yellow cards.",
    ("fantasy_player_round_stats", "RC"):                "Red cards.",
    ("fantasy_player_round_stats", "OG"):                "Own goals.",
    ("fantasy_player_round_stats", "PW"):                "Penalty won.",
    ("fantasy_player_round_stats", "PC"):                "Penalty conceded.",
    ("fantasy_player_round_stats", "CS"):                "Clean sheet (1/0).",
    ("fantasy_player_round_stats", "GS"):                "Goals scored.",
    ("fantasy_player_round_stats", "GC"):                "Goals conceded by player's team while on the pitch.",
    ("fantasy_player_round_stats", "PS"):                "Penalty saved.",
    ("fantasy_player_round_stats", "T"):                 "Tackles. **Midfielder Volume Bonus trigger (≥4.5/90 per D-1-Doc §2).**",
    ("fantasy_player_round_stats", "CC"):                "Chances created. **Midfielder Volume Bonus trigger (≥1.5/90).**",
    ("fantasy_player_round_stats", "ST"):                "Shots on target.",
    ("fantasy_player_round_stats", "FK"):                "Free kicks taken.",
    ("fantasy_player_round_stats", "S"):                 "Saves (GK).",
    ("fantasy_player_round_stats", "SB"):                "**Scouting Bonus eligibility flag (1/0).** True when player scored >4 pts with <5% ownership.",
}

# ─── Style helpers
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SUB_FILL    = PatternFill("solid", fgColor="DDEBF7")
WRAP        = Alignment(wrap_text=True, vertical="top")
BORDER      = Border(*(Side(style="thin", color="BFBFBF") for _ in range(4)))

def safe_sample(s: pd.Series) -> str:
    try:
        nn = s.dropna()
        if not len(nn): return ""
        v = nn.iloc[0]
        out = str(v)
        return out[:80] + ("…" if len(out) > 80 else "")
    except Exception:
        return ""

def autosize(ws, widths: dict[int, int]):
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

# ─── Build workbook
wb = Workbook()
wb.remove(wb.active)

# Overview sheet
ov = wb.create_sheet("Overview")
ov.append(["WC26 Fantasy Intelligence Engine — Data Dictionary"])
ov["A1"].font = Font(bold=True, size=14)
ov.append([])
ov.append(["Companion docs: METRICS_MAP.md (per-metric × source), REFRESH.md (per-table refresh cadence)."])
ov.append([])
ov.append(["Table", "Rows", "Cols", "Notebook", "Refresh", "Description"])
hdr_row = 5
for c in range(1, 7):
    ov.cell(row=hdr_row, column=c).font = HEADER_FONT
    ov.cell(row=hdr_row, column=c).fill = HEADER_FILL

# Walk every processed parquet, build overview + per-sheet
parquets = sorted(PROCESSED.glob("*.parquet"))
print(f"found {len(parquets)} parquet files")
for pq in parquets:
    table = pq.stem
    df = pd.read_parquet(pq)
    notes = TABLE_NOTES.get(table, {})

    ov.append([
        table, len(df), len(df.columns),
        notes.get("notebook", ""),
        notes.get("refresh", ""),
        notes.get("desc", ""),
    ])

    sheet_name = table[:31]
    ws = wb.create_sheet(sheet_name)

    # Top header block
    ws.append([table])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([f"{len(df):,} rows × {len(df.columns)} cols  |  notebook {notes.get('notebook','?')}  |  refresh: {notes.get('refresh','?')}"])
    ws["A2"].font = Font(italic=True)
    ws.append([notes.get("desc", "")])
    ws["A3"].alignment = WRAP
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    ws.append([])
    ws.append(["column", "dtype", "non-null", "null", "% null", "sample", "description"])
    for c in range(1, 8):
        cell = ws.cell(row=5, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for col in df.columns:
        s = df[col]
        nn = int(s.notna().sum())
        null = len(df) - nn
        pct = round(100 * null / len(df), 1) if len(df) else 0
        desc = COL_DESC.get((table, col), "")
        ws.append([col, str(s.dtype), nn, null, pct, safe_sample(s), desc])
        for c in range(1, 8):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP
            ws.cell(row=ws.max_row, column=c).border = BORDER

    autosize(ws, {1: 32, 2: 14, 3: 11, 4: 8, 5: 8, 6: 40, 7: 70})
    ws.freeze_panes = "A6"

autosize(ov, {1: 38, 2: 9, 3: 7, 4: 11, 5: 14, 6: 80})
ov.freeze_panes = "A6"

wb.save(OUT)
print(f"wrote {OUT.name}  ({OUT.stat().st_size / 1024:.0f} KB)  with {len(wb.sheetnames)} sheets")
